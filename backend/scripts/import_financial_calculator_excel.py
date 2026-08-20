"""导入快速财务计算器 Excel 基础数据。

该脚本用于把《工艺消耗表格.xlsx》的“组成 / 消耗 / 产出”三张表结构化写入
现有财务模型配置体系。脚本保持可重复执行：同一批基础数据按固定编码更新，
节点消耗、路线节点和路线产出会以当前 Excel 内容整体替换。
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.exceptions import AppException
from app.core.database import SessionLocal
from app.models.process_config import (
    ProcessCalculationImportBatch,
    ProcessConsumable,
    ProcessMaterial,
    ProcessNode,
    ProcessProduct,
    ProcessPublicService,
    ProcessRoute,
)
from app.models.user import User
from app.repositories.process_config_repository import (
    ProcessLibraryRepository,
    ProcessNodeRepository,
    ProcessRouteRepository,
)
from app.schemas.process_config import (
    ProcessCalculationOutputPayload,
    ProcessCalculationOutputReplacePayload,
    ProcessConsumableCreateWithPrices,
    ProcessConsumableUpdateWithPrices,
    ProcessMaterialCompositionPayload,
    ProcessMaterialCompositionReplacePayload,
    ProcessMaterialCreateWithPrices,
    ProcessMaterialUpdateWithPrices,
    ProcessNodeConsumablePayload,
    ProcessNodeCreateWithChildren,
    ProcessNodeOutputPayload,
    ProcessNodePublicServicePayload,
    ProcessNodeUpdateWithChildren,
    ProcessProductCreateWithPrices,
    ProcessProductUpdateWithPrices,
    ProcessPublicServiceCreateWithPrices,
    ProcessPublicServiceUpdateWithPrices,
    ProcessRouteCreateWithNodes,
    ProcessRouteNodePayload,
    ProcessRouteUpdateWithNodes,
)
from app.services.process_config_service import ProcessConfigService

logger = logging.getLogger(__name__)

EXCEL_IMPORT_TYPE = "financial_calculator_excel"
MATERIAL_CODE = "M1"
LEGACY_MATERIAL_CODE = "MAT_BM_EXCEL"
LEGACY_ROUTE_CODE = "ROUTE_EXCEL_BM_RECYCLE"
SOURCE_VERSION = "V1"
LEGACY_SOURCE_VERSIONS = ("excel-20260709", SOURCE_VERSION)

CATEGORY_TO_OUTPUT_TYPE = {
    "产品": "product",
    "废固": "solid_waste",
    "废水": "wastewater",
    "废气": "waste_gas",
}

TARGET_OUTPUT_CATEGORY_BY_NAME = {
    "精制硫酸锂溶液": "li",
    "工业级碳酸锂": "li",
    "电池级碳酸锂": "li",
    "电池级硫酸锰": "mn",
    "粗制碳酸锰": "mn",
    "电池级硫酸钴": "co",
    "电池级硫酸镍": "ni",
    "粗制海绵铜": "cu",
    "粗制硫化铜": "cu",
    "硫酸铜溶液": "cu",
    "金属铜产品": "cu",
    "石墨渣产品": "graphite",
}

ROUTE_OPTION_BY_NODE_NAME = {
    "混酸焙烧": ("root_leach_path", "mixed_acid_roasting"),
    "酸浸": ("root_leach_path", "acid_leaching"),
    "水浸渣酸浸": ("water_leach_residue_path", "residue_acid_leaching"),
    "氧化镍钴渣产品": ("water_leach_residue_path", "oxide_ni_co_slag_product"),
    "石墨渣废固": ("graphite_handling", "untreated_solid_waste"),
    "石墨干燥": ("graphite_handling", "drying"),
    "除铜-铁粉法": ("copper_removal", "iron_powder"),
    "除铜-硫化法": ("copper_removal", "sulfide"),
    "除铜-萃取法": ("copper_removal", "extraction"),
    "树脂除氟(PREL)": ("lithium_purification", "prel_resin_defluorination"),
    "硫酸锂深度化学除杂": ("lithium_purification", "deep_chemical_purification"),
    "MHP沉淀": ("nickel_cobalt_manganese_path", "mhp_precipitation"),
    "P204萃取": ("nickel_cobalt_manganese_path", "p204_extraction"),
    "工业级碳酸锂干燥包装": ("lithium_carbonate_finish", "industrial_drying"),
    "碳化热析": ("lithium_carbonate_finish", "carbonation_thermal_decomposition"),
    "锰溶液除杂": ("manganese_product_path", "manganese_solution_purification"),
    "碳酸锰沉淀": ("manganese_product_path", "manganese_carbonate_precipitation"),
    "BC196共萃": ("ni_co_separation", "bc196_coextraction"),
    "C272萃取钴": ("ni_co_separation", "c272_cobalt_extraction"),
    "P507萃取钴": ("ni_co_separation", "p507_cobalt_extraction"),
    "P507萃镍镁": ("ni_co_separation", "p507_cobalt_extraction"),
}

NODE_WASTE_OUTPUTS_BY_OUTPUT_NAME = {
    "废弃石墨渣": ("石墨渣废固",),
    "水解除杂渣": ("水解除杂",),
    "硫酸锂深度除杂渣": ("硫酸锂深度化学除杂",),
    "锰溶液除杂渣": ("锰溶液除杂",),
    "P204反铁废水": ("P204萃取",),
    "P204洗氯废水": ("P204萃取",),
    "NaCl废水(PMN)": ("锰溶液除杂",),
    "NaCl废水(SXMN)": ("C272萃锰",),
    "除氟废水": ("除氟树脂",),
    "树脂除氟废水(PREL)": ("树脂除氟(PREL)",),
    "树脂除钙镁废水(PREL)": ("树脂除钙镁(PREL)",),
    "C272萃锰反铁废水": ("C272萃锰",),
    "BC196共萃反铁废水": ("BC196共萃",),
    "BC196萃镍反铁废水": ("BC196萃镍",),
    "C272萃钴反铁废水": ("C272萃取钴",),
    "C272萃镁反铁废水": ("C272萃取镁",),
    "P507萃镍镁反铁废水": ("P507萃镍镁",),
    "P507萃钴反铁废水": ("P507萃取钴",),
}
OBSOLETE_OUTPUT_NAMES = {"水解除杂渣(铁粉法)"}
NODE_PRODUCT_OUTPUTS_BY_OUTPUT_NAME = {
    "石墨渣产品": ("石墨干燥",),
}
ACTIVATED_CARBON_OUTPUT_NAME = "活性炭渣"
ACTIVATED_CARBON_RESOURCE_NAME = "活性炭"
ACTIVATED_CARBON_EFFECTIVE_RATIO = Decimal("0.9")

# 节点编码按技术路线树的“首次出现层级”命名；同一工艺节点在不同分支复用同一编码。
NODE_CODE_BY_NAME = {
    "混酸焙烧": "A1",
    "酸浸": "A2",
    "混酸焙烧尾气处理": "B1",
    "水浸": "B2",
    "石墨干燥": "B3",
    "除铜-铁粉法": "B4",
    "除铜-硫化法": "B5",
    "除铜-萃取法": "B6",
    "石墨渣废固": "B7",
    "水浸渣酸浸": "B8",
    "氧化镍钴渣产品": "B9",
    "水浸液化学除杂": "C1",
    "除铜-萃取&电积": "C2",
    "水解除杂": "C3",
    "树脂除氟(PREL)": "D1",
    "硫酸锂深度化学除杂": "D2",
    "除氟树脂": "D3",
    "树脂除钙镁(PREL)": "E1",
    "碳酸锂沉淀": "E2",
    "MHP沉淀": "E3",
    "P204萃取": "E4",
    "工业级碳酸锂干燥包装": "F1",
    "碳化热析": "F2",
    "硫酸钠蒸发&低温结晶(PREL)": "F3",
    "硫酸钠蒸发&低温结晶(MHP)": "F4",
    "锰溶液除杂": "F5",
    "BC196共萃": "F6",
    "P507萃取钴": "F7",
    "P507萃镍镁": "F8",
    "硫酸钠蒸发&低温结晶(SX)": "F9",
    "碳酸锰沉淀": "F10",
    "电池级碳酸锂干燥包装": "G1",
    "硫酸钠干燥包装(PREL)": "G2",
    "硫酸钠干燥包装(MHP)": "G3",
    "C272萃锰": "G4",
    "C272萃取钴": "G5",
    "BC196萃镍": "G6",
    "C272萃取镁": "G7",
    "硫酸钠干燥包装(SX)": "G8",
    "硫酸锰蒸发结晶": "H1",
    "硫酸钴蒸发结晶": "H2",
    "硫酸镍蒸发结晶": "H3",
    "电池级硫酸锰干燥包装": "I1",
    "电池级硫酸钴干燥包装": "I2",
    "电池级硫酸镍干燥包装": "I3",
}
NODE_SORT_ORDER_BY_NAME = {
    name: index for index, name in enumerate(NODE_CODE_BY_NAME, start=1)
}
SYNTHETIC_ROUTE_NODE_NAMES = ("石墨渣废固", "水浸渣酸浸", "氧化镍钴渣产品", "碳酸锰沉淀")
MUTUALLY_EXCLUSIVE_ROUTE_NODE_GROUPS = (
    ("根级浸出工艺", ("混酸焙烧", "酸浸")),
    ("水浸滤饼处理", ("水浸渣酸浸", "氧化镍钴渣产品")),
    ("石墨渣处理", ("石墨渣废固", "石墨干燥")),
    ("锂液除杂工艺", ("树脂除氟(PREL)", "硫酸锂深度化学除杂")),
    ("碳酸锂产品工艺", ("工业级碳酸锂干燥包装", "碳化热析")),
    ("镍钴锰主线", ("MHP沉淀", "P204萃取")),
    ("锰产品工艺", ("锰溶液除杂", "碳酸锰沉淀")),
    ("钴分离工艺", ("BC196共萃", "C272萃取钴", "P507萃取钴", "P507萃镍镁")),
)

SYNTHETIC_OUTPUT_ROWS: tuple[dict[str, Any], ...] = (
    {
        "row_index": 10039,
        "code": "P39",
        "legacy_code": "OUT_SYNTHETIC_PRODUCT_OXIDE_NI_CO_SLAG",
        "output_type": "product",
        "output_name": "氧化镍钴渣",
        "spec": None,
        "recovery_rate": Decimal("1"),
        "balance_weight": Decimal("0"),
        "unit": "t/t-BM",
        "output_ratio": Decimal("1"),
        "formula_type": "fixed",
        "expression": None,
        "scale_param": {
            "source": "updated_route_drawing",
            "synthetic": True,
            "note": "水浸滤饼直接作为产品，产出系数待项目数据补齐",
        },
    },
)

GENERATED_ROUTE_CODE_PATTERN = re.compile(r"^[A-Z]\d+(?:-[A-Z]\d+)+$")


@dataclass(frozen=True)
class CompositionRow:
    element_code: str
    element_name: str
    content_ratio: Decimal
    unit: str


@dataclass(frozen=True)
class ResourceRow:
    row_index: int
    code: str
    legacy_code: str
    name: str
    spec: str | None
    balance_weight: Decimal
    unit: str
    resource_type: str


@dataclass(frozen=True)
class CoefficientCell:
    amount: Decimal
    formula_type: str
    expression: str | None
    source_row: int
    source_col: int
    source_cell: str


@dataclass(frozen=True)
class NodeRow:
    index: int
    code: str
    name: str
    source_col: int


@dataclass(frozen=True)
class OutputRow:
    row_index: int
    code: str
    legacy_code: str
    output_type: str
    output_name: str
    spec: str | None
    recovery_rate: Decimal
    balance_weight: Decimal
    unit: str
    output_ratio: Decimal
    formula_type: str
    expression: str | None
    scale_param: dict[str, Any]
    sort_order: int


@dataclass(frozen=True)
class RouteDefinition:
    code: str
    name: str
    output: OutputRow
    node_names: list[str]
    node_codes: list[str]


@dataclass(frozen=True)
class ParsedWorkbook:
    source_path: Path
    compositions: list[CompositionRow]
    consumables: list[ResourceRow]
    public_services: list[ResourceRow]
    nodes: list[NodeRow]
    coefficients: dict[tuple[int, int], CoefficientCell]
    outputs: list[OutputRow]
    formula_count: int
    error_count: int


def append_synthetic_outputs(outputs: list[OutputRow]) -> None:
    """补图纸中存在、但历史 Excel 产出表暂未维护的产品叶子。"""

    existing_names = {row.output_name for row in outputs}
    existing_codes = {row.code for row in outputs}
    next_sort_order = max((row.sort_order for row in outputs), default=0) + 1
    for row in SYNTHETIC_OUTPUT_ROWS:
        if row["output_name"] in existing_names:
            continue
        code = str(row["code"])
        if code in existing_codes:
            code = f"P{len(outputs) + 1}"
        outputs.append(
            OutputRow(
                row_index=int(row["row_index"]),
                code=code,
                legacy_code=str(row["legacy_code"]),
                output_type=str(row["output_type"]),
                output_name=str(row["output_name"]),
                spec=row["spec"],
                recovery_rate=row["recovery_rate"],
                balance_weight=row["balance_weight"],
                unit=str(row["unit"]),
                output_ratio=row["output_ratio"],
                formula_type=str(row["formula_type"]),
                expression=row["expression"],
                scale_param=dict(row["scale_param"]),
                sort_order=next_sort_order,
            )
        )
        existing_names.add(str(row["output_name"]))
        existing_codes.add(code)
        next_sort_order += 1


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    )


def utc_now_naive() -> datetime:
    """返回 naive UTC 时间，兼容项目当前数据库字段写法。"""

    return datetime.now(UTC).replace(tzinfo=None)


def is_generated_route(route: ProcessRoute) -> bool:
    """判断是否为本导入脚本生成的路线，避免误清理用户手工维护路线。"""

    if route.code == LEGACY_ROUTE_CODE:
        return True
    return route.version in LEGACY_SOURCE_VERSIONS and bool(
        GENERATED_ROUTE_CODE_PATTERN.fullmatch(route.code)
    )


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "/":
        return None
    return text


def to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    text = clean_text(value)
    if text is None:
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default


def is_numeric(value: Any) -> bool:
    if value is None:
        return False
    try:
        Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return False
    return True


def has_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def is_error_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("#")


def build_spec(value: Any, unit: Any = None) -> str | None:
    spec_value = clean_text(value)
    spec_unit = clean_text(unit)
    if spec_value and spec_unit:
        return f"{spec_value} {spec_unit}"
    return spec_value or spec_unit


def make_source_cell(row_index: int, col_index: int) -> str:
    return f"{row_index}:{col_index}"


def find_default_excel_path() -> Path:
    download_dir = Path("E:/download")
    candidates = [
        path
        for path in download_dir.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("%")
    ]
    if not candidates:
        raise FileNotFoundError("未在 E:/download 找到可导入的 xlsx 文件")
    return min(candidates, key=lambda path: len(path.name))


def parse_coefficient_cell(
    formula_cell: Cell,
    value_cell: Cell,
    row_index: int,
    col_index: int,
) -> CoefficientCell | None:
    raw_formula = formula_cell.value
    raw_value = value_cell.value
    if clean_text(raw_formula) is None and clean_text(raw_value) is None:
        return None

    expression: str | None = None
    formula_type = "fixed"
    if has_formula(formula_cell):
        formula_type = "expression"
        expression = str(raw_formula)
    elif not is_numeric(raw_formula) and clean_text(raw_formula):
        formula_type = "expression"
        expression = str(raw_formula).strip()

    amount = to_decimal(raw_value)
    return CoefficientCell(
        amount=amount,
        formula_type=formula_type,
        expression=expression,
        source_row=row_index,
        source_col=col_index,
        source_cell=make_source_cell(row_index, col_index),
    )


def parse_workbook(source_path: Path) -> ParsedWorkbook:
    formula_wb = load_workbook(source_path, data_only=False)
    value_wb = load_workbook(source_path, data_only=True)

    formula_composition_ws = formula_wb.worksheets[0]
    value_composition_ws = value_wb.worksheets[0]
    formula_consumption_ws = formula_wb.worksheets[1]
    value_consumption_ws = value_wb.worksheets[1]
    formula_output_ws = formula_wb.worksheets[2]
    value_output_ws = value_wb.worksheets[2]

    compositions: list[CompositionRow] = []
    for row_index in range(2, value_composition_ws.max_row + 1):
        element = clean_text(value_composition_ws.cell(row=row_index, column=2).value)
        if not element:
            continue
        compositions.append(
            CompositionRow(
                element_code=element,
                element_name=element,
                content_ratio=to_decimal(
                    value_composition_ws.cell(row=row_index, column=3).value
                ),
                unit="ratio",
            )
        )

    nodes: list[NodeRow] = []
    for col_index in range(5, value_consumption_ws.max_column + 1):
        node_name = clean_text(value_consumption_ws.cell(row=1, column=col_index).value)
        if not node_name:
            continue
        node_code = NODE_CODE_BY_NAME.get(node_name)
        if node_code is None:
            raise AppException(message=f"未配置工艺节点编码映射: {node_name}")
        nodes.append(
            NodeRow(
                index=NODE_SORT_ORDER_BY_NAME[node_name],
                code=node_code,
                name=node_name,
                source_col=col_index,
            )
        )

    consumables: list[ResourceRow] = []
    public_services: list[ResourceRow] = []
    coefficients: dict[tuple[int, int], CoefficientCell] = {}
    current_section = "consumable"

    for row_index in range(2, value_consumption_ws.max_row + 1):
        name = clean_text(value_consumption_ws.cell(row=row_index, column=1).value)
        if not name:
            continue
        if name == "公用工程":
            current_section = "public_service"
            continue

        row_list = public_services if current_section == "public_service" else consumables
        code_prefix = "U" if current_section == "public_service" else "C"
        legacy_code_prefix = "PS_EXCEL" if current_section == "public_service" else "CONS_EXCEL"
        resource_index = len(row_list) + 1
        resource = ResourceRow(
            row_index=row_index,
            code=f"{code_prefix}{resource_index}",
            legacy_code=f"{legacy_code_prefix}_{resource_index:03d}",
            name=name,
            spec=clean_text(value_consumption_ws.cell(row=row_index, column=2).value),
            balance_weight=to_decimal(
                value_consumption_ws.cell(row=row_index, column=3).value
            ),
            unit=clean_text(value_consumption_ws.cell(row=row_index, column=4).value) or "",
            resource_type=current_section,
        )
        row_list.append(resource)

        for node in nodes:
            coefficient = parse_coefficient_cell(
                formula_consumption_ws.cell(row=row_index, column=node.source_col),
                value_consumption_ws.cell(row=row_index, column=node.source_col),
                row_index,
                node.source_col,
            )
            if coefficient is not None:
                coefficients[(row_index, node.source_col)] = coefficient

    existing_node_names = {node.name for node in nodes}
    for node_name in SYNTHETIC_ROUTE_NODE_NAMES:
        if node_name in existing_node_names:
            continue
        nodes.append(
            NodeRow(
                index=NODE_SORT_ORDER_BY_NAME[node_name],
                code=NODE_CODE_BY_NAME[node_name],
                name=node_name,
                source_col=0,
            )
        )
    nodes.sort(key=lambda node: node.index)

    outputs: list[OutputRow] = []
    current_category: str | None = None
    category_index: dict[str, int] = {}
    for row_index in range(1, value_output_ws.max_row + 1):
        first_col_value = clean_text(value_output_ws.cell(row=row_index, column=1).value)
        if first_col_value in CATEGORY_TO_OUTPUT_TYPE:
            current_category = first_col_value
            category_index.setdefault(current_category, 0)
            continue
        if not current_category or not first_col_value:
            continue

        output_type = CATEGORY_TO_OUTPUT_TYPE[current_category]
        category_index[current_category] += 1
        type_code = re.sub(r"[^A-Z_]", "_", output_type.upper())
        value_ratio_cell = value_output_ws.cell(row=row_index, column=7)
        formula_ratio_cell = formula_output_ws.cell(row=row_index, column=7)
        value_balance_cell = value_output_ws.cell(row=row_index, column=5)
        formula_balance_cell = formula_output_ws.cell(row=row_index, column=5)

        formula_type = "fixed"
        expression: str | None = None
        if has_formula(formula_ratio_cell):
            formula_type = "expression"
            expression = str(formula_ratio_cell.value)
        elif not is_numeric(formula_ratio_cell.value) and clean_text(formula_ratio_cell.value):
            formula_type = "expression"
            expression = str(formula_ratio_cell.value).strip()

        scale_param: dict[str, Any] = {
            "source_sheet": formula_output_ws.title,
            "source_row": row_index,
            "source_cell": make_source_cell(row_index, 7),
        }
        if has_formula(formula_balance_cell):
            scale_param["balance_weight_expression"] = str(formula_balance_cell.value)
        elif (
            not is_numeric(formula_balance_cell.value)
            and clean_text(formula_balance_cell.value)
        ):
            scale_param["balance_weight_expression"] = str(formula_balance_cell.value).strip()

        outputs.append(
            OutputRow(
                row_index=row_index,
                code=f"P{len(outputs) + 1}",
                legacy_code=f"OUT_EXCEL_{type_code}_{category_index[current_category]:03d}",
                output_type=output_type,
                output_name=first_col_value,
                spec=build_spec(
                    value_output_ws.cell(row=row_index, column=2).value,
                    value_output_ws.cell(row=row_index, column=3).value,
                ),
                recovery_rate=to_decimal(value_output_ws.cell(row=row_index, column=4).value),
                balance_weight=to_decimal(value_balance_cell.value),
                unit=clean_text(value_output_ws.cell(row=row_index, column=6).value) or "",
                output_ratio=to_decimal(value_ratio_cell.value),
                formula_type=formula_type,
                expression=expression,
                scale_param=scale_param,
                sort_order=len(outputs) + 1,
            )
        )
    append_synthetic_outputs(outputs)

    formula_count = 0
    error_count = 0
    for worksheet in formula_wb.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if has_formula(cell):
                    formula_count += 1
                if is_error_value(cell.value):
                    error_count += 1

    return ParsedWorkbook(
        source_path=source_path,
        compositions=compositions,
        consumables=consumables,
        public_services=public_services,
        nodes=nodes,
        coefficients=coefficients,
        outputs=outputs,
        formula_count=formula_count,
        error_count=error_count,
    )


def copy_source_file(source_path: Path) -> Path:
    target_dir = BACKEND_DIR / "storage" / "process-calculation-imports"
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    target_path = target_dir / f"{timestamp}_{source_path.name}"
    shutil.copy2(source_path, target_path)
    return target_path


def get_operator(db) -> User:
    operator = db.query(User).filter(User.id == 1, User.is_deleted.is_(False)).first()
    if operator is None:
        raise AppException(message="未找到导入操作人 admin(id=1)，无法写入审计字段")
    return operator


def find_active_library_by_codes(
    repo: ProcessLibraryRepository,
    code: str,
    legacy_code: str | None = None,
) -> Any | None:
    """按新旧编码查找基础库记录，支持导入编码体系升级时原地改名。"""

    for candidate_code in [code, legacy_code]:
        if not candidate_code:
            continue
        item = repo.get_by_code(candidate_code)
        if item is not None and not item.is_deleted:
            return item
    return None


def create_import_batch(
    db,
    source_path: Path,
    stored_path: Path,
    operator_id: int,
) -> ProcessCalculationImportBatch:
    batch = ProcessCalculationImportBatch(
        file_name=source_path.name,
        file_path=str(stored_path),
        import_type=EXCEL_IMPORT_TYPE,
        status="pending",
        success_count=0,
        failed_count=0,
        created_by=operator_id,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch


def finish_import_batch(
    db,
    batch_id: int,
    status: str,
    success_count: int,
    failed_count: int = 0,
    error_message: str | None = None,
) -> None:
    batch = db.get(ProcessCalculationImportBatch, batch_id)
    if batch is None:
        return
    batch.status = status
    batch.success_count = success_count
    batch.failed_count = failed_count
    batch.error_message = error_message
    db.commit()


def upsert_material(
    db,
    service: ProcessConfigService,
    operator: User,
    parsed: ParsedWorkbook,
) -> dict[str, Any]:
    repo = ProcessLibraryRepository(db, ProcessMaterial, "material")
    existing = find_active_library_by_codes(repo, MATERIAL_CODE, LEGACY_MATERIAL_CODE)
    payload_data = {
        "code": MATERIAL_CODE,
        "name": "黑粉BM",
        "type": "battery_black_mass",
        "unit": "t-BM",
        "status": "enabled",
        "description": "动力电池黑粉原料，维护元素组成用于快速测算。",
    }
    if existing and not existing.is_deleted:
        update_payload = ProcessMaterialUpdateWithPrices(**payload_data)
        return service.update_library("material", existing.id, update_payload, operator)
    create_payload = ProcessMaterialCreateWithPrices(**payload_data)
    return service.create_library("material", create_payload, operator)


def upsert_consumable(
    db,
    service: ProcessConfigService,
    operator: User,
    row: ResourceRow,
) -> dict[str, Any]:
    repo = ProcessLibraryRepository(db, ProcessConsumable, "consumable")
    existing = find_active_library_by_codes(repo, row.code, row.legacy_code)
    payload_data = {
        "code": row.code,
        "name": row.name,
        "type": "chemical",
        "unit": row.unit or "t/t-BM",
        "status": "enabled",
        "description": json.dumps(
            {
                "category": "chemical",
                "excel_row": row.row_index,
                "spec": row.spec,
            },
            ensure_ascii=False,
        ),
    }
    if existing and not existing.is_deleted:
        return service.update_library(
            "consumable",
            existing.id,
            ProcessConsumableUpdateWithPrices(**payload_data),
            operator,
        )
    return service.create_library(
        "consumable",
        ProcessConsumableCreateWithPrices(**payload_data),
        operator,
    )


def upsert_public_service(
    db,
    service: ProcessConfigService,
    operator: User,
    row: ResourceRow,
) -> dict[str, Any]:
    repo = ProcessLibraryRepository(db, ProcessPublicService, "public_service")
    existing = find_active_library_by_codes(repo, row.code, row.legacy_code)
    payload_data = {
        "code": row.code,
        "name": row.name,
        "type": "utility",
        "unit": row.unit,
        "status": "enabled",
        "description": json.dumps(
            {"category": "public_service", "excel_row": row.row_index},
            ensure_ascii=False,
        ),
    }
    if existing and not existing.is_deleted:
        return service.update_library(
            "public_service",
            existing.id,
            ProcessPublicServiceUpdateWithPrices(**payload_data),
            operator,
        )
    return service.create_library(
        "public_service",
        ProcessPublicServiceCreateWithPrices(**payload_data),
        operator,
    )


def upsert_product_output(
    db,
    service: ProcessConfigService,
    operator: User,
    row: OutputRow,
) -> dict[str, Any]:
    repo = ProcessLibraryRepository(db, ProcessProduct, "product")
    existing = find_active_library_by_codes(repo, row.code, row.legacy_code)
    payload_data = {
        "code": row.code,
        "name": row.output_name,
        "type": row.output_type,
        "output_type": row.output_type,
        "target_output_category": TARGET_OUTPUT_CATEGORY_BY_NAME.get(row.output_name),
        "is_product_form": row.output_type in {"product", "byproduct"},
        "spec": row.spec,
        "unit": row.unit or "t/t-BM",
        "status": "enabled",
        "treatment_cost": Decimal("0"),
        "description": json.dumps(
            {
                "category": row.output_type,
                "excel_row": row.row_index,
                "recovery_rate": str(row.recovery_rate),
                "output_ratio": str(row.output_ratio),
            },
            ensure_ascii=False,
        ),
    }
    if existing and not existing.is_deleted:
        return service.update_library(
            "product",
            existing.id,
            ProcessProductUpdateWithPrices(**payload_data),
            operator,
        )
    return service.create_library(
        "product",
        ProcessProductCreateWithPrices(**payload_data),
        operator,
    )


def replace_material_compositions(
    service: ProcessConfigService,
    material_id: int,
    parsed: ParsedWorkbook,
    operator: User,
) -> int:
    payload = ProcessMaterialCompositionReplacePayload(
        items=[
            ProcessMaterialCompositionPayload(
                element_code=row.element_code,
                element_name=row.element_name,
                content_ratio=row.content_ratio,
                unit=row.unit,
                remark="原料元素组成",
            )
            for row in parsed.compositions
        ]
    )
    service.replace_material_compositions(material_id, payload, operator)
    return len(payload.items)


def parse_spec_number(spec: str | None, default: Decimal = Decimal("1")) -> Decimal:
    if not spec:
        return default
    match = re.search(r"-?\d+(?:\.\d+)?", spec)
    if not match:
        return default
    return to_decimal(match.group(0), default)


def build_node_waste_output_payloads(
    node: NodeRow,
    parsed: ParsedWorkbook,
    product_ids: dict[int, int],
    batch_id: int,
) -> list[ProcessNodeOutputPayload]:
    """把技术路线树中的紫色三废产出绑定到实际产生它的工艺节点。"""

    outputs_by_name = {row.output_name: row for row in parsed.outputs}
    payloads: list[ProcessNodeOutputPayload] = []

    for output_name, node_names in NODE_WASTE_OUTPUTS_BY_OUTPUT_NAME.items():
        if node.name not in node_names:
            continue
        output = outputs_by_name.get(output_name)
        if output is None:
            continue
        payloads.append(
            ProcessNodeOutputPayload(
                product_id=product_ids[output.row_index],
                output_type=output.output_type,
                output_per_ton=output.output_ratio,
                formula_type=output.formula_type,
                expression=output.expression,
                scale_param={
                    **output.scale_param,
                    "source_sheet": "产出",
                    "node_output_role": "waste_treatment",
                    "binding_rule": "route_tree",
                    "node_name": node.name,
                    "route_condition": {},
                },
                source_template_id=batch_id,
                balance_weight=output.balance_weight,
                treatment_cost=Decimal("0"),
                unit=output.unit,
                is_main_product=False,
                sort_order=len(payloads) + 1,
                remark="节点三废产出",
            )
        )

    for output_name, node_names in NODE_PRODUCT_OUTPUTS_BY_OUTPUT_NAME.items():
        if node.name not in node_names:
            continue
        output = outputs_by_name.get(output_name)
        if output is None:
            continue
        payloads.append(
            ProcessNodeOutputPayload(
                product_id=product_ids[output.row_index],
                output_type=output.output_type,
                output_per_ton=output.output_ratio,
                formula_type=output.formula_type,
                expression=output.expression,
                scale_param={
                    **output.scale_param,
                    "source_sheet": "产出",
                    "node_output_role": "product_output",
                    "binding_rule": "route_tree",
                    "node_name": node.name,
                    "route_condition": {},
                },
                source_template_id=batch_id,
                balance_weight=output.balance_weight,
                treatment_cost=Decimal("0"),
                unit=output.unit,
                is_main_product=True,
                sort_order=len(payloads) + 1,
                remark="节点产品产出",
            )
        )

    activated_carbon_output = outputs_by_name.get(ACTIVATED_CARBON_OUTPUT_NAME)
    activated_carbon_row = next(
        (row for row in parsed.consumables if row.name == ACTIVATED_CARBON_RESOURCE_NAME),
        None,
    )
    if activated_carbon_output is not None and activated_carbon_row is not None:
        coefficient = parsed.coefficients.get((activated_carbon_row.row_index, node.source_col))
        if coefficient is not None and coefficient.amount > 0:
            spec_number = parse_spec_number(activated_carbon_output.spec, Decimal("0.5"))
            denominator = ACTIVATED_CARBON_EFFECTIVE_RATIO * spec_number
            output_per_ton = coefficient.amount / denominator if denominator else Decimal("0")
            payloads.append(
                ProcessNodeOutputPayload(
                    product_id=product_ids[activated_carbon_output.row_index],
                    output_type=activated_carbon_output.output_type,
                    output_per_ton=output_per_ton,
                    formula_type=activated_carbon_output.formula_type,
                    expression=activated_carbon_output.expression,
                    scale_param={
                        **activated_carbon_output.scale_param,
                        "source_sheet": "产出",
                        "node_output_role": "waste_treatment",
                        "binding_rule": "activated_carbon_consumption",
                        "node_name": node.name,
                        "source_consumable_code": activated_carbon_row.code,
                        "source_consumable_name": activated_carbon_row.name,
                        "source_consumption_cell": coefficient.source_cell,
                        "source_consumption_amount": str(coefficient.amount),
                    },
                    source_template_id=batch_id,
                    balance_weight=activated_carbon_output.balance_weight,
                    treatment_cost=Decimal("0"),
                    unit=activated_carbon_output.unit,
                    is_main_product=False,
                    sort_order=len(payloads) + 1,
                    remark="节点三废产出",
                )
            )

    return payloads


def build_product_route_paths() -> dict[str, list[list[str]]]:
    """按技术线路树生成产品目标路线。

    路线以图纸结构为准：少数仅用于表达互斥分支的逻辑节点允许没有 Excel 消耗列，
    例如“石墨渣废固”和“碳酸锰沉淀”。
    """

    lithium_routes = {
        "精制硫酸锂溶液": [
            ["混酸焙烧", "水浸", "水浸液化学除杂", "树脂除氟(PREL)", "树脂除钙镁(PREL)"]
        ],
        "工业级碳酸锂": [
            ["混酸焙烧", "水浸", "水浸液化学除杂", "硫酸锂深度化学除杂", "碳酸锂沉淀", "工业级碳酸锂干燥包装"]
        ],
        "电池级碳酸锂": [
            ["混酸焙烧", "水浸", "水浸液化学除杂", "硫酸锂深度化学除杂", "碳酸锂沉淀", "碳化热析", "电池级碳酸锂干燥包装"]
        ],
        "硫酸钠产品(PREL）": [
            [
                "混酸焙烧",
                "水浸",
                "水浸液化学除杂",
                "硫酸锂深度化学除杂",
                "碳酸锂沉淀",
                "硫酸钠蒸发&低温结晶(PREL)",
                "硫酸钠干燥包装(PREL)",
            ]
        ],
    }

    # 图纸中“酸浸”既是黑粉根路径的互斥选项，也出现在混酸焙烧后水浸渣的下游处理。
    # 这里将混酸焙烧后的下游产出挂在“混酸焙烧 -> 水浸”前缀下，避免把根级 acid_leaching
    # 选项误并入同一候选路线，同时保留 Li 与 Ni/Co/Mn/Cu 等叶子产出同属 A1 方案的结构。
    acid_leach_graphite_waste_prefix = ["酸浸", "石墨渣废固"]
    acid_leach_graphite_product_prefix = ["酸浸", "石墨干燥"]
    acid_leach_prefixes = [
        acid_leach_graphite_waste_prefix,
        acid_leach_graphite_product_prefix,
    ]
    mixed_acid_residue_graphite_waste_prefix = ["混酸焙烧", "水浸", "水浸渣酸浸", "石墨渣废固"]
    mixed_acid_residue_graphite_product_prefix = ["混酸焙烧", "水浸", "水浸渣酸浸", "石墨干燥"]
    mixed_acid_residue_prefixes = [
        mixed_acid_residue_graphite_waste_prefix,
        mixed_acid_residue_graphite_product_prefix,
    ]
    downstream_prefixes = [*acid_leach_prefixes, *mixed_acid_residue_prefixes]
    copper_product_tails = {
        "粗制海绵铜": [["除铜-铁粉法"]],
        "粗制硫化铜": [["除铜-硫化法"]],
        "硫酸铜溶液": [["除铜-萃取法"]],
        "金属铜产品": [["除铜-萃取法", "除铜-萃取&电积"]],
    }
    copper_removal_tails = [
        ["除铜-铁粉法"],
        ["除铜-硫化法"],
        ["除铜-萃取法"],
    ]
    downstream_tails = {
        "MHP滤饼产品": [["水解除杂", "除氟树脂", "MHP沉淀"]],
        "硫酸钠产品(MHP)": [["水解除杂", "除氟树脂", "MHP沉淀", "硫酸钠蒸发&低温结晶(MHP)", "硫酸钠干燥包装(MHP)"]],
        "电池级硫酸锰": [["水解除杂", "除氟树脂", "P204萃取", "锰溶液除杂", "C272萃锰", "硫酸锰蒸发结晶", "电池级硫酸锰干燥包装"]],
        "粗制碳酸锰": [["水解除杂", "除氟树脂", "P204萃取", "碳酸锰沉淀"]],
        "硫酸镍钴溶液": [["水解除杂", "除氟树脂", "P204萃取", "BC196共萃"]],
        "电池级硫酸钴": [
            ["水解除杂", "除氟树脂", "P204萃取", "BC196共萃", "硫酸钴蒸发结晶", "电池级硫酸钴干燥包装"],
            ["水解除杂", "除氟树脂", "P204萃取", "C272萃取钴", "硫酸钴蒸发结晶", "电池级硫酸钴干燥包装"],
            ["水解除杂", "除氟树脂", "P204萃取", "P507萃取钴", "硫酸钴蒸发结晶", "电池级硫酸钴干燥包装"],
        ],
        "电池级硫酸镍": [
            ["水解除杂", "除氟树脂", "P204萃取", "BC196共萃", "BC196萃镍", "硫酸镍蒸发结晶", "电池级硫酸镍干燥包装"],
            ["水解除杂", "除氟树脂", "P204萃取", "P507萃镍镁", "C272萃取镁", "硫酸镍蒸发结晶", "电池级硫酸镍干燥包装"],
        ],
        "硫酸钠产品(SX)": [["水解除杂", "除氟树脂", "P204萃取", "硫酸钠蒸发&低温结晶(SX)", "硫酸钠干燥包装(SX)"]],
    }

    route_paths: dict[str, list[list[str]]] = dict(lithium_routes)
    route_paths["氧化镍钴渣"] = [["混酸焙烧", "水浸", "氧化镍钴渣产品"]]
    route_paths["石墨渣产品"] = [
        acid_leach_graphite_product_prefix,
        mixed_acid_residue_graphite_product_prefix,
    ]

    for output_name, tails in copper_product_tails.items():
        route_paths[output_name] = [prefix + tail for prefix in downstream_prefixes for tail in tails]

    for output_name, tails in downstream_tails.items():
        route_paths[output_name] = [
            prefix + copper_tail + tail
            for prefix in downstream_prefixes
            for copper_tail in copper_removal_tails
            for tail in tails
        ]

    return {
        output_name: [normalize_route_path(path) for path in paths]
        for output_name, paths in route_paths.items()
    }


def normalize_route_path(path: list[str]) -> list[str]:
    """按新图纸规则补必达节点，并拒绝同层级互斥工艺同时出现在一条路线中。"""

    normalized = list(path)
    if "混酸焙烧" in normalized and "混酸焙烧尾气处理" not in normalized:
        insert_at = normalized.index("混酸焙烧") + 1
        normalized.insert(insert_at, "混酸焙烧尾气处理")
    if "混酸焙烧" in normalized and "水浸" not in normalized:
        raise AppException(message="混酸焙烧路线必须同时包含混酸焙烧尾气处理和水浸节点")
    for group_name, node_names in MUTUALLY_EXCLUSIVE_ROUTE_NODE_GROUPS:
        selected_nodes = [node_name for node_name in node_names if node_name in normalized]
        if len(selected_nodes) > 1:
            raise AppException(
                message=f"{group_name}属于同一互斥工艺选项组，不能同时生成在一条工艺路线中: {', '.join(selected_nodes)}"
            )
    return normalized


def build_route_node_params(
    node_name: str,
    route_node_names: list[str] | None = None,
    node_index: int | None = None,
) -> str | None:
    """为路线树预览提供图纸分叉语义，不改变成本计算的线性节点集合。"""

    parent_acid_leach_code: str | None = None
    has_later_copper_removal = False
    if route_node_names is not None and node_index is not None:
        for upstream_node_name in reversed(route_node_names[:node_index]):
            if upstream_node_name == "水浸渣酸浸":
                parent_acid_leach_code = "B8"
                break
            if upstream_node_name == "酸浸":
                parent_acid_leach_code = "A2"
                break
        has_later_copper_removal = any(
            downstream_node_name in {"除铜-铁粉法", "除铜-硫化法", "除铜-萃取法"}
            for downstream_node_name in route_node_names[node_index + 1 :]
        )

    params_by_node_name: dict[str, dict[str, Any]] = {
        "混酸焙烧尾气处理": {
            "parent_node_code": "A1",
            "parallel_group_code": "mixed_acid_primary_outputs",
            "branch_role": "tail_gas_treatment",
            "continues_route": False,
        },
        "水浸": {
            "parent_node_code": "A1",
            "parallel_group_code": "mixed_acid_primary_outputs",
            "branch_role": "water_leach",
            "continues_route": True,
        },
        "水浸渣酸浸": {
            "parent_node_code": "B2",
            "parallel_group_code": "water_leach_residue_outputs",
            "branch_role": "residue_acid_leaching",
            "continues_route": True,
        },
        "氧化镍钴渣产品": {
            "parent_node_code": "B2",
            "parallel_group_code": "water_leach_residue_outputs",
            "branch_role": "oxide_ni_co_slag_product",
            "continues_route": True,
        },
    }
    if parent_acid_leach_code and node_name in {"石墨干燥", "石墨渣废固"}:
        params_by_node_name[node_name] = {
            "parent_node_code": parent_acid_leach_code,
            "parallel_group_code": "acid_leach_primary_outputs",
            "branch_role": "graphite_drying" if node_name == "石墨干燥" else "graphite_solid_waste",
            "continues_route": not has_later_copper_removal,
        }
    if parent_acid_leach_code and node_name in {"除铜-铁粉法", "除铜-硫化法", "除铜-萃取法"}:
        params_by_node_name[node_name] = {
            "parent_node_code": parent_acid_leach_code,
            "parallel_group_code": "acid_leach_primary_outputs",
            "branch_role": "copper_removal",
            "continues_route": True,
        }
    params = params_by_node_name.get(node_name)
    if not params:
        return None
    return json.dumps(params, ensure_ascii=False, separators=(",", ":"))


def build_route_definitions(parsed: ParsedWorkbook) -> list[RouteDefinition]:
    outputs_by_name = {
        row.output_name: row
        for row in parsed.outputs
        if row.output_type == "product"
    }
    paths_by_output = build_product_route_paths()
    definitions: list[RouteDefinition] = []
    used_codes: set[str] = set()

    for output_name, paths in paths_by_output.items():
        output = outputs_by_name.get(output_name)
        if output is None:
            logger.warning("产品路线未生成，产出Sheet缺少目标产品: output_name=%s", output_name)
            continue
        for path in paths:
            node_codes = []
            for node_name in path:
                code = NODE_CODE_BY_NAME.get(node_name)
                if code is None:
                    raise AppException(message=f"路线节点缺少编码映射: output={output_name} node={node_name}")
                node_codes.append(code)
            route_code = "-".join(node_codes)
            if route_code in used_codes:
                raise AppException(message=f"路线编码重复，请检查技术线路树映射: {route_code}")
            used_codes.add(route_code)
            definitions.append(
                RouteDefinition(
                    code=route_code,
                    name=f"{output_name}路线（{route_code}）",
                    output=output,
                    node_names=path,
                    node_codes=node_codes,
                )
            )

    unmapped_outputs = sorted(set(outputs_by_name) - set(paths_by_output))
    if unmapped_outputs:
        logger.warning("存在未映射的产品产出，未生成目标路线: outputs=%s", unmapped_outputs)
    return definitions


def soft_delete_generated_routes(
    db,
    service: ProcessConfigService,
    operator: User,
) -> int:
    """重建路线前先软删除本脚本生成的活动路线，释放旧节点引用。"""

    deleted_count = 0
    for route in (
        db.query(ProcessRoute)
        .filter(ProcessRoute.is_deleted.is_(False))
        .all()
    ):
        if not is_generated_route(route):
            continue
        service.delete_route(route.id, operator)
        deleted_count += 1

    return deleted_count


def make_archived_node_code(node: ProcessNode) -> str:
    """生成不会占用新编码体系的归档节点编码。"""

    code = f"ARCHIVED_NODE_{node.id}_{node.code}"
    return code[:100]


def archive_conflicting_generated_nodes(
    db,
    operator: User,
    desired_code_by_name: dict[str, str],
) -> int:
    """归档旧导入节点，解决编码重排后 code 全局唯一冲突。"""

    desired_codes = set(desired_code_by_name.values())
    desired_name_by_code = {code: name for name, code in desired_code_by_name.items()}
    desired_pairs = {(code, name) for name, code in desired_code_by_name.items()}
    node_repo = ProcessNodeRepository(db)
    candidates: dict[int, ProcessNode] = {}

    for node in db.query(ProcessNode).filter(ProcessNode.code.in_(desired_codes)).all():
        if desired_name_by_code.get(node.code) != node.name:
            candidates[node.id] = node

    for node in (
        db.query(ProcessNode)
        .filter(ProcessNode.version.in_(LEGACY_SOURCE_VERSIONS))
        .all()
    ):
        if (node.code, node.name) in desired_pairs:
            continue
        if node.is_deleted and node.code not in desired_codes:
            continue
        candidates[node.id] = node

    archived_count = 0
    for node in candidates.values():
        if node_repo.count_route_references(node.id) > 0:
            raise AppException(message=f"节点编码重排前仍被活动路线引用: node_id={node.id} code={node.code}")
        old_code = node.code
        node.code = make_archived_node_code(node)
        node_repo.soft_delete(node)
        node.updated_by = operator.id
        archived_count += 1
        logger.info(
            "旧导入节点已归档: node_id=%s old_code=%s archived_code=%s name=%s",
            node.id,
            old_code,
            node.code,
            node.name,
        )

    if archived_count:
        db.commit()
    return archived_count


def upsert_nodes(
    db,
    service: ProcessConfigService,
    operator: User,
    parsed: ParsedWorkbook,
    consumable_ids: dict[int, int],
    public_service_ids: dict[int, int],
    product_ids: dict[int, int],
    batch_id: int,
) -> dict[str, int]:
    node_repo = ProcessNodeRepository(db)
    node_ids_by_name: dict[str, int] = {}
    consumable_rows = {row.row_index: row for row in parsed.consumables}
    public_service_rows = {row.row_index: row for row in parsed.public_services}

    for node in parsed.nodes:
        consumable_payloads: list[ProcessNodeConsumablePayload] = []
        public_service_payloads: list[ProcessNodePublicServicePayload] = []

        for row_index, row in consumable_rows.items():
            coefficient = parsed.coefficients.get((row_index, node.source_col))
            if coefficient is None:
                continue
            consumable_payloads.append(
                ProcessNodeConsumablePayload(
                    consumable_id=consumable_ids[row_index],
                    amount_per_ton=coefficient.amount,
                    formula_type=coefficient.formula_type,
                    amount_per_ton_bm=coefficient.amount,
                    expression=coefficient.expression,
                    scale_param={
                        "source_sheet": "消耗",
                        "source_row": coefficient.source_row,
                        "source_col": coefficient.source_col,
                        "source_cell": coefficient.source_cell,
                    },
                    source_template_id=batch_id,
                    balance_weight=row.balance_weight,
                    unit=row.unit,
                    sort_order=len(consumable_payloads) + 1,
                )
            )

        for row_index, row in public_service_rows.items():
            coefficient = parsed.coefficients.get((row_index, node.source_col))
            if coefficient is None:
                continue
            public_service_payloads.append(
                ProcessNodePublicServicePayload(
                    public_service_id=public_service_ids[row_index],
                    amount_per_ton=coefficient.amount,
                    formula_type=coefficient.formula_type,
                    amount_per_ton_bm=coefficient.amount,
                    expression=coefficient.expression,
                    scale_param={
                        "source_sheet": "消耗",
                        "source_row": coefficient.source_row,
                        "source_col": coefficient.source_col,
                        "source_cell": coefficient.source_cell,
                    },
                    source_template_id=batch_id,
                    balance_weight=row.balance_weight,
                    unit=row.unit,
                    sort_order=len(public_service_payloads) + 1,
                )
            )

        existing = node_repo.get_by_code(node.code)
        output_payloads = build_node_waste_output_payloads(node, parsed, product_ids, batch_id)
        payload_data = {
            "code": node.code,
            "name": node.name,
            "node_type": "hydrometallurgy",
            "version": SOURCE_VERSION,
            "status": "enabled",
            "description": (
                f"工艺节点消耗系数配置，数据列 {node.source_col}"
                if node.source_col > 0
                else "图纸逻辑节点，无Excel消耗系数"
            ),
            "sort_order": node.index,
            "consumables": consumable_payloads,
            "public_services": public_service_payloads,
            "material_inputs": [],
            "outputs": output_payloads,
            "equipment": [],
        }

        if existing and existing.is_deleted:
            existing.is_deleted = False
            existing.deleted_at = None
            existing.status = "enabled"
            db.commit()
            db.refresh(existing)

        if existing:
            result = service.update_node(
                existing.id,
                ProcessNodeUpdateWithChildren(**payload_data),
                operator,
            )
        else:
            result = service.create_node(
                ProcessNodeCreateWithChildren(**payload_data),
                operator,
            )
        node_ids_by_name[node.name] = int(result["id"])
    return node_ids_by_name


def upsert_route_definition(
    db,
    service: ProcessConfigService,
    operator: User,
    material_id: int,
    parsed: ParsedWorkbook,
    node_ids_by_name: dict[str, int],
    product_ids: dict[int, int],
    route_definition: RouteDefinition,
) -> int:
    route_repo = ProcessRouteRepository(db)
    existing = route_repo.get_by_code(route_definition.code)
    route_nodes = [
        ProcessRouteNodePayload(
            node_id=node_ids_by_name[node_name],
            sort_order=index,
            option_group_code=(ROUTE_OPTION_BY_NODE_NAME.get(node_name) or (None, None))[0],
            option_code=(ROUTE_OPTION_BY_NODE_NAME.get(node_name) or (None, None))[1],
            node_params_json=build_route_node_params(
                node_name,
                route_definition.node_names,
                index - 1,
            ),
        )
        for index, node_name in enumerate(route_definition.node_names, start=1)
    ]
    payload_data = {
        "code": route_definition.code,
        "name": route_definition.name,
        "route_type": "recycle",
        "input_material_id": material_id,
        "final_product_id": product_ids[route_definition.output.row_index],
        "version": SOURCE_VERSION,
        "status": "enabled",
        "description": f"目标产出：{route_definition.output.output_name}",
        "nodes": route_nodes,
    }

    if existing and existing.is_deleted:
        existing.is_deleted = False
        existing.deleted_at = None
        existing.status = "enabled"
        db.commit()
        db.refresh(existing)

    if existing:
        result = service.update_route(
            existing.id,
            ProcessRouteUpdateWithNodes(**payload_data),
            operator,
        )
    else:
        result = service.create_route(
            ProcessRouteCreateWithNodes(**payload_data),
            operator,
        )
    if "id" in result:
        return int(result["id"])
    return int(result["route"]["id"])


def replace_route_output(
    service: ProcessConfigService,
    route_id: int,
    parsed: ParsedWorkbook,
    product_ids: dict[int, int],
    route_definition: RouteDefinition,
    operator: User,
) -> int:
    row = route_definition.output
    item = ProcessCalculationOutputPayload(
        output_type=row.output_type,
        product_id=product_ids.get(row.row_index),
        output_name=row.output_name,
        spec=row.spec,
        recovery_rate=row.recovery_rate,
        balance_weight=row.balance_weight,
        unit=row.unit,
        output_ratio=row.output_ratio,
        treatment_cost=Decimal("0"),
        formula_type=row.formula_type,
        expression=row.expression,
        scale_param={
            **row.scale_param,
            "route_code": route_definition.code,
            "route_nodes": route_definition.node_names,
            "route_node_codes": route_definition.node_codes,
        },
        sort_order=1,
        remark="路线目标产出系数",
    )

    payload = ProcessCalculationOutputReplacePayload(items=[item])
    service.replace_route_calculation_outputs(route_id, payload, operator)
    return 1


def import_parsed_workbook(parsed: ParsedWorkbook) -> int:
    stored_path = copy_source_file(parsed.source_path)
    success_count = 0

    db = SessionLocal()
    batch_id: int | None = None
    try:
        operator = get_operator(db)
        batch = create_import_batch(db, parsed.source_path, stored_path, operator.id)
        batch_id = batch.id
        service = ProcessConfigService(db)

        material = upsert_material(db, service, operator, parsed)
        material_id = int(material["id"])
        success_count += 1
        success_count += replace_material_compositions(service, material_id, parsed, operator)

        consumable_ids: dict[int, int] = {}
        for row in parsed.consumables:
            result = upsert_consumable(db, service, operator, row)
            consumable_ids[row.row_index] = int(result["id"])
            success_count += 1

        public_service_ids: dict[int, int] = {}
        for row in parsed.public_services:
            result = upsert_public_service(db, service, operator, row)
            public_service_ids[row.row_index] = int(result["id"])
            success_count += 1

        product_ids: dict[int, int] = {}
        for row in parsed.outputs:
            if row.output_name in OBSOLETE_OUTPUT_NAMES:
                continue
            result = upsert_product_output(db, service, operator, row)
            product_ids[row.row_index] = int(result["id"])
            success_count += 1

        route_definitions = build_route_definitions(parsed)
        success_count += soft_delete_generated_routes(db, service, operator)
        success_count += archive_conflicting_generated_nodes(db, operator, NODE_CODE_BY_NAME)

        node_ids = upsert_nodes(
            db,
            service,
            operator,
            parsed,
            consumable_ids,
            public_service_ids,
            product_ids,
            batch_id,
        )
        success_count += len(node_ids)

        route_ids = []
        for route_definition in route_definitions:
            route_id = upsert_route_definition(
                db,
                service,
                operator,
                material_id,
                parsed,
                node_ids,
                product_ids,
                route_definition,
            )
            route_ids.append(route_id)
            success_count += 1
            success_count += replace_route_output(
                service,
                route_id,
                parsed,
                product_ids,
                route_definition,
                operator,
            )

        success_count += len(parsed.coefficients)
        finish_import_batch(db, batch_id, "success", success_count)
        logger.info(
            "Excel导入完成 batch_id=%s material_id=%s route_count=%s success_count=%s formulas=%s errors=%s",
            batch_id,
            material_id,
            len(route_ids),
            success_count,
            parsed.formula_count,
            parsed.error_count,
        )
        return batch_id
    except Exception as exc:
        db.rollback()
        if batch_id is not None:
            finish_import_batch(
                db,
                batch_id,
                "failed",
                success_count,
                failed_count=1,
                error_message=str(exc)[:2000],
            )
        logger.exception("Excel导入失败 batch_id=%s source=%s", batch_id, parsed.source_path)
        raise
    finally:
        db.close()


def main() -> int:
    configure_logging()
    parser = argparse.ArgumentParser(description="导入快速财务计算器 Excel 基础数据")
    parser.add_argument(
        "excel_path",
        nargs="?",
        type=Path,
        default=None,
        help="Excel 文件路径，默认自动查找 E:/download 下的 xlsx",
    )
    args = parser.parse_args()

    source_path = args.excel_path or find_default_excel_path()
    source_path = source_path.resolve()
    logger.info("开始解析Excel source=%s", source_path)
    parsed = parse_workbook(source_path)
    logger.info(
        "Excel解析完成 compositions=%s consumables=%s public_services=%s nodes=%s coefficients=%s outputs=%s formulas=%s errors=%s",
        len(parsed.compositions),
        len(parsed.consumables),
        len(parsed.public_services),
        len(parsed.nodes),
        len(parsed.coefficients),
        len(parsed.outputs),
        parsed.formula_count,
        parsed.error_count,
    )
    import_parsed_workbook(parsed)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
