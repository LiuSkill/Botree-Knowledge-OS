"""Process configuration library service tests."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR))

from app.api.deps import get_current_user  # noqa: E402
from app.core.database import get_db  # noqa: E402
from app.core.exceptions import AppException  # noqa: E402
from app.models import (  # noqa: E402
    Base,
    ProcessNode,
    ProcessNodeConsumable,
    ProcessNodeEquipment,
    ProcessNodeMaterialInput,
    ProcessNodeOutput,
    ProcessNodePublicService,
    ProcessRoute,
    ProcessRouteNode,
    Role,
    User,
)
from app.schemas.process_config import (  # noqa: E402
    ProcessConsumableCreateWithPrices,
    ProcessLibraryRegionPricePayload,
    ProcessLibraryStatusUpdate,
    ProcessMaterialCreateWithPrices,
    ProcessMaterialUpdateWithPrices,
    ProcessNodeCreateWithChildren,
    ProcessNodeUpdateWithChildren,
    ProcessProductCreateWithPrices,
    ProcessPublicServiceCreateWithPrices,
    ProcessRouteCreateWithNodes,
    ProcessRouteNodeReorderPayload,
    ProcessRouteUpdateWithNodes,
    ProcessRouteVersionCreatePayload,
)
from app.services.process_config_excel_service import ProcessConfigExcelService  # noqa: E402
from app.services.process_config_service import ProcessConfigService  # noqa: E402
from main import app  # noqa: E402


@pytest.fixture()
def db_session() -> Session:
    """Create an isolated in-memory database session."""

    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    with session_factory() as session:
        yield session
    engine.dispose()


@pytest.fixture()
def api_db_session() -> Session:
    """Create a TestClient-friendly in-memory database session."""

    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    session = session_factory()
    try:
        yield session
    finally:
        app.dependency_overrides.clear()
        session.close()
        engine.dispose()


def seed_operator(db: Session) -> User:
    """Seed an operator for operation logs."""

    operator = User(username="admin", password_hash="x", real_name="Admin", status="enabled")
    db.add(operator)
    db.commit()
    db.refresh(operator)
    return operator


def region_prices(asia_price: str = "10.5") -> list[ProcessLibraryRegionPricePayload]:
    """Build the three required region prices."""

    return [
        ProcessLibraryRegionPricePayload(region_code="asia", unit_price=Decimal(asia_price), unit="kg"),
        ProcessLibraryRegionPricePayload(region_code="europe", unit_price=Decimal("20.5"), unit="kg"),
        ProcessLibraryRegionPricePayload(region_code="americas", unit_price=Decimal("30.5"), unit="kg"),
    ]


def workbook_bytes(workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_material(service: ProcessConfigService, operator: User, code: str = "M001") -> dict:
    return service.create_library(
        "material",
        ProcessMaterialCreateWithPrices(
            code=code,
            name="Black Mass",
            type="battery",
            description="Cathode feed",
            unit="kg",
            status="enabled",
            region_prices=region_prices(),
        ),
        operator,
    )


def create_product(service: ProcessConfigService, operator: User, code: str = "P001") -> dict:
    return service.create_library(
        "product",
        ProcessProductCreateWithPrices(code=code, name="Nickel Salt", type="metal", description="Final product", unit="kg", status="enabled"),
        operator,
    )


def create_consumable(service: ProcessConfigService, operator: User, code: str = "C001") -> dict:
    return service.create_library(
        "consumable",
        ProcessConsumableCreateWithPrices(code=code, name="Sulfuric Acid", type="chemical", description="Acid", unit="kg", status="enabled"),
        operator,
    )


def create_public_service(service: ProcessConfigService, operator: User, code: str = "S001") -> dict:
    return service.create_library(
        "public_service",
        ProcessPublicServiceCreateWithPrices(code=code, name="Electricity", type="utility", description="Power", unit="kWh", status="enabled"),
        operator,
    )


def node_payload(
    material_id: int,
    product_id: int,
    consumable_id: int,
    public_service_id: int,
    *,
    code: str = "N001",
    name: str = "Leaching",
    status: str = "enabled",
) -> ProcessNodeCreateWithChildren:
    """Build a complete node payload with all child sections."""

    return ProcessNodeCreateWithChildren(
        code=code,
        name=name,
        node_type="hydrometallurgy",
        staff=Decimal("2"),
        area=Decimal("120"),
        description="Hydrometallurgy leaching",
        status=status,
        version="1.0",
        sort_order=1,
        remark="node remark",
        material_inputs=[
            {"material_id": material_id, "amount_per_ton": Decimal("1.5"), "unit": "t", "sort_order": 1},
        ],
        consumables=[
            {"consumable_id": consumable_id, "amount_per_ton": Decimal("0.2"), "unit": "kg", "sort_order": 1},
        ],
        public_services=[
            {"public_service_id": public_service_id, "amount_per_ton": Decimal("10"), "unit": "kWh", "sort_order": 1},
        ],
        equipment=[
            {
                "equipment_name": "Reactor",
                "equipment_type": "tank",
                "quantity": Decimal("2"),
                "investment_amount": Decimal("100000"),
                "currency": "CNY",
                "sort_order": 1,
                "remark": "main equipment",
            },
        ],
        outputs=[
            {"product_id": product_id, "output_per_ton": Decimal("0.8"), "unit": "kg", "is_main_product": True, "sort_order": 1},
        ],
    )


def route_payload(
    material_id: int,
    product_id: int,
    node_ids: list[int],
    *,
    code: str = "R001",
    name: str = "Route 001",
    status: str = "enabled",
) -> ProcessRouteCreateWithNodes:
    """Build a complete route payload with node chain."""

    return ProcessRouteCreateWithNodes(
        code=code,
        name=name,
        input_material_id=material_id,
        final_product_id=product_id,
        version="1.0",
        description="Route description",
        status=status,
        sort_order=1,
        remark="route remark",
        nodes=[
            {
                "node_id": node_id,
                "sort_order": index,
                "node_params_json": {"recovery": f"{90 + index}%"},
                "remark": f"step {index}",
            }
            for index, node_id in enumerate(node_ids, start=1)
        ],
    )


def test_process_config_api_routes_are_registered_and_protected(api_db_session: Session) -> None:
    """Process config APIs are mounted and protected by backend permission checks."""

    operator = seed_operator(api_db_session)

    def override_db() -> Session:
        return api_db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: operator

    paths = {route.path for route in app.routes}
    assert "/api/process-config/materials" in paths
    assert "/api/process-config/materials/template" in paths
    assert "/api/process-config/materials/export" in paths
    assert "/api/process-config/materials/import" in paths
    assert "/api/process-config/products" in paths
    assert "/api/process-config/products/template" in paths
    assert "/api/process-config/products/export" in paths
    assert "/api/process-config/products/import" in paths
    assert "/api/process-config/consumables" in paths
    assert "/api/process-config/consumables/template" in paths
    assert "/api/process-config/consumables/export" in paths
    assert "/api/process-config/consumables/import" in paths
    assert "/api/process-config/public-services" in paths
    assert "/api/process-config/public-services/template" in paths
    assert "/api/process-config/public-services/export" in paths
    assert "/api/process-config/public-services/import" in paths
    assert "/api/process-config/nodes" in paths
    assert "/api/process-config/nodes/{node_id}" in paths
    assert "/api/process-config/nodes/template" in paths
    assert "/api/process-config/nodes/export" in paths
    assert "/api/process-config/nodes/import" in paths
    assert "/api/process-config/routes" in paths
    assert "/api/process-config/routes/{route_id}" in paths
    assert "/api/process-config/routes/template" in paths
    assert "/api/process-config/routes/export" in paths
    assert "/api/process-config/routes/import" in paths
    assert "/api/process-config/routes/{route_id}/copy" in paths
    assert "/api/process-config/routes/{route_id}/versions" in paths
    assert "/api/process-config/options/materials" in paths

    response = TestClient(app).get("/api/process-config/materials")

    assert response.status_code == 403
    assert response.json()["message"] == "无权执行该操作"


def test_material_api_crud_roundtrip(api_db_session: Session) -> None:
    """Admin can use the material CRUD API end-to-end."""

    operator = seed_operator(api_db_session)
    admin_role = Role(name="Admin", code="admin", enabled=True)
    operator.roles = [admin_role]
    api_db_session.add(admin_role)
    api_db_session.commit()

    def override_db() -> Session:
        return api_db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: operator
    client = TestClient(app)

    create_response = client.post(
        "/api/process-config/materials",
        json={
            "code": "M-API",
            "name": "API Material",
            "type": "battery",
            "description": "API feed",
            "unit": "kg",
            "status": "enabled",
            "region_prices": [
                {"region_code": "asia", "unit_price": "1.1", "unit": "kg"},
                {"region_code": "europe", "unit_price": "2.2", "unit": "kg"},
                {"region_code": "americas", "unit_price": "3.3", "unit": "kg"},
            ],
        },
    )

    assert create_response.status_code == 200
    created = create_response.json()["data"]
    assert created["code"] == "M-API"
    assert [price["currency"] for price in created["region_prices"]] == ["CNY", "EUR", "USD"]

    list_response = client.get("/api/process-config/materials", params={"keyword": "API", "status": "enabled"})
    assert list_response.status_code == 200
    assert list_response.json()["data"]["total"] == 1

    detail_response = client.get(f"/api/process-config/materials/{created['id']}")
    assert detail_response.status_code == 200
    assert detail_response.json()["data"]["name"] == "API Material"

    status_response = client.patch(f"/api/process-config/materials/{created['id']}/status", json={"status": "disabled"})
    assert status_response.status_code == 200
    assert status_response.json()["data"]["status"] == "disabled"

    delete_response = client.delete(f"/api/process-config/materials/{created['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["data"] == {"deleted": True}


def test_material_excel_api_roundtrip(api_db_session: Session) -> None:
    """Admin can download the material template, import Excel data, and export current records."""

    operator = seed_operator(api_db_session)
    admin_role = Role(name="Admin", code="admin", enabled=True)
    operator.roles = [admin_role]
    api_db_session.add(admin_role)
    api_db_session.commit()

    def override_db() -> Session:
        return api_db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: operator
    client = TestClient(app)

    template_response = client.get("/api/process-config/materials/template")
    assert template_response.status_code == 200
    assert template_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(template_response.content))
    workbook["原料库"].append(
        ["M-EXCEL-API", "Excel Material", "battery", "From API import", "kg", "enabled", 1, "remark", "1.1", "2.2", "3.3"]
    )

    import_response = client.post(
        "/api/process-config/materials/import",
        files={
            "file": (
                "materials.xlsx",
                workbook_bytes(workbook),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    import_data = import_response.json()["data"]
    assert import_data["module"] == "materials"
    assert import_data["imported_count"] == 1
    assert import_data["imported_codes"] == ["M-EXCEL-API"]

    export_response = client.get("/api/process-config/materials/export", params={"keyword": "M-EXCEL-API"})
    assert export_response.status_code == 200
    export_workbook = load_workbook(BytesIO(export_response.content), data_only=True)
    rows = list(export_workbook["原料库"].iter_rows(values_only=True))
    assert rows[1][0] == "M-EXCEL-API"
    assert rows[1][1] == "Excel Material"


def test_node_api_crud_roundtrip(api_db_session: Session) -> None:
    """Admin can use the node CRUD API end-to-end."""

    operator = seed_operator(api_db_session)
    admin_role = Role(name="Admin", code="admin", enabled=True)
    operator.roles = [admin_role]
    api_db_session.add(admin_role)
    api_db_session.commit()

    service = ProcessConfigService(api_db_session)
    material = create_material(service, operator, code="M-NODE-API")
    product = create_product(service, operator, code="P-NODE-API")
    consumable = create_consumable(service, operator, code="C-NODE-API")
    public_service = create_public_service(service, operator, code="S-NODE-API")

    def override_db() -> Session:
        return api_db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: operator
    client = TestClient(app)
    payload = node_payload(
        material["id"],
        product["id"],
        consumable["id"],
        public_service["id"],
        code="N-API",
        name="API Node",
    ).model_dump(mode="json")

    create_response = client.post("/api/process-config/nodes", json=payload)
    assert create_response.status_code == 200
    created = create_response.json()["data"]
    assert created["code"] == "N-API"
    assert len(created["outputs"]) == 1

    list_response = client.get("/api/process-config/nodes", params={"keyword": "API", "node_type": "hydrometallurgy", "status": "enabled"})
    assert list_response.status_code == 200
    assert list_response.json()["data"]["total"] == 1

    detail_response = client.get(f"/api/process-config/nodes/{created['id']}")
    assert detail_response.status_code == 200
    assert detail_response.json()["data"]["equipment"][0]["equipment_name"] == "Reactor"

    payload["name"] = "Updated API Node"
    payload["outputs"][0]["output_per_ton"] = "0.95"
    update_response = client.put(f"/api/process-config/nodes/{created['id']}", json=payload)
    assert update_response.status_code == 200
    updated = update_response.json()["data"]
    assert updated["name"] == "Updated API Node"
    assert Decimal(updated["outputs"][0]["output_per_ton"]) == Decimal("0.95")

    delete_response = client.delete(f"/api/process-config/nodes/{created['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["data"] == {"deleted": True}


def test_route_api_crud_roundtrip(api_db_session: Session) -> None:
    """Admin can use the route CRUD, copy, and version APIs end-to-end."""

    operator = seed_operator(api_db_session)
    admin_role = Role(name="Admin", code="admin", enabled=True)
    operator.roles = [admin_role]
    api_db_session.add(admin_role)
    api_db_session.commit()

    service = ProcessConfigService(api_db_session)
    material = create_material(service, operator, code="M-ROUTE-API")
    product = create_product(service, operator, code="P-ROUTE-API")
    consumable = create_consumable(service, operator, code="C-ROUTE-API")
    public_service = create_public_service(service, operator, code="S-ROUTE-API")
    first_node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-API-1", name="Route Node 1"),
        operator,
    )
    second_node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-API-2", name="Route Node 2"),
        operator,
    )

    def override_db() -> Session:
        return api_db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: operator
    client = TestClient(app)

    create_response = client.post(
        "/api/process-config/routes",
        json=route_payload(material["id"], product["id"], [first_node["id"]], code="R-API", name="API Route").model_dump(mode="json"),
    )
    assert create_response.status_code == 200
    created = create_response.json()["data"]
    assert created["route"]["code"] == "R-API"
    assert created["nodes"][0]["node"]["code"] == "N-ROUTE-API-1"
    assert json.loads(created["nodes"][0]["node_params_json"])["recovery"] == "91%"

    list_response = client.get("/api/process-config/routes", params={"keyword": "API", "status": "enabled"})
    assert list_response.status_code == 200
    assert list_response.json()["data"]["total"] == 1

    add_node_response = client.post(
        f"/api/process-config/routes/{created['route']['id']}/nodes",
        json={"node_id": second_node["id"], "sort_order": 2, "node_params_json": {"recovery": "93%"}, "remark": "step 2"},
    )
    assert add_node_response.status_code == 200
    added_route_node = add_node_response.json()["data"]
    assert added_route_node["sort_order"] == 2

    reorder_response = client.put(
        f"/api/process-config/routes/{created['route']['id']}/nodes/reorder",
        json={
            "items": [
                {"route_node_id": created["nodes"][0]["id"], "sort_order": 2},
                {"route_node_id": added_route_node["id"], "sort_order": 1},
            ]
        },
    )
    assert reorder_response.status_code == 200
    reordered = reorder_response.json()["data"]
    assert [row["sort_order"] for row in reordered] == [1, 2]
    assert reordered[0]["node_id"] == second_node["id"]

    update_response = client.put(
        f"/api/process-config/routes/{created['route']['id']}",
        json=ProcessRouteUpdateWithNodes(
            name="Updated API Route",
            nodes=[
                {"node_id": second_node["id"], "sort_order": 1, "node_params_json": {"recovery": "93%"}, "remark": "step 2"},
                {"node_id": first_node["id"], "sort_order": 2, "node_params_json": {"recovery": "91%"}, "remark": "step 1"},
            ],
        ).model_dump(mode="json", exclude_unset=True),
    )
    assert update_response.status_code == 200
    updated = update_response.json()["data"]
    assert updated["route"]["name"] == "Updated API Route"
    assert updated["nodes"][0]["node"]["code"] == "N-ROUTE-API-2"

    detail_response = client.get(f"/api/process-config/routes/{created['route']['id']}")
    assert detail_response.status_code == 200
    assert detail_response.json()["data"]["final_product"]["code"] == "P-ROUTE-API"
    assert len(detail_response.json()["data"]["nodes"]) == 2

    version_response = client.post(
        f"/api/process-config/routes/{created['route']['id']}/versions",
        json=ProcessRouteVersionCreatePayload(change_log="baseline").model_dump(mode="json", exclude_none=True),
    )
    assert version_response.status_code == 200
    version = version_response.json()["data"]
    assert version["version_no"] == 1
    assert json.loads(version["snapshot_json"])["route"]["code"] == "R-API"

    versions_response = client.get(f"/api/process-config/routes/{created['route']['id']}/versions")
    assert versions_response.status_code == 200
    assert versions_response.json()["data"][0]["version_no"] == 1

    copy_response = client.post(f"/api/process-config/routes/{created['route']['id']}/copy")
    assert copy_response.status_code == 200
    copied = copy_response.json()["data"]
    assert copied["route"]["status"] == "draft"
    assert copied["route"]["code"].startswith("R-API_COPY")
    assert len(copied["nodes"]) == 2

    delete_route_node_response = client.delete(f"/api/process-config/routes/{created['route']['id']}/nodes/{updated['nodes'][1]['id']}")
    assert delete_route_node_response.status_code == 200
    assert delete_route_node_response.json()["data"] == {"deleted": True}

    delete_response = client.delete(f"/api/process-config/routes/{created['route']['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["data"] == {"deleted": True}


def test_library_crud_persists_region_prices_and_options(db_session: Session) -> None:
    """Creating and updating a library item persists the three region prices."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)

    created = create_material(service, operator)

    assert created["code"] == "M001"
    assert [price["region_code"] for price in created["region_prices"]] == ["asia", "europe", "americas"]
    assert [price["currency"] for price in created["region_prices"]] == ["CNY", "EUR", "USD"]

    page = service.list_library("material", keyword="feed", status="enabled", page=1, page_size=10)
    assert page["total"] == 1
    assert Decimal(page["items"][0]["region_prices"][0]["unit_price"]) == Decimal("10.5")

    updated = service.update_library(
        "material",
        created["id"],
        ProcessMaterialUpdateWithPrices(name="Updated Black Mass", region_prices=region_prices(asia_price="99.9")),
        operator,
    )

    assert updated["name"] == "Updated Black Mass"
    assert Decimal(updated["region_prices"][0]["unit_price"]) == Decimal("99.9")

    disabled = service.update_status("material", created["id"], ProcessLibraryStatusUpdate(status="disabled"), operator)

    assert disabled["status"] == "disabled"
    assert service.list_options("material") == []

    service.delete_library("material", created["id"], operator)
    with pytest.raises(AppException, match="数据不存在"):
        service.get_library("material", created["id"])


def test_node_crud_persists_child_sections(db_session: Session) -> None:
    """Creating, updating and deleting a node keeps all child sections consistent."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    material = create_material(service, operator, code="M-NODE")
    product = create_product(service, operator, code="P-NODE")
    consumable = create_consumable(service, operator, code="C-NODE")
    public_service = create_public_service(service, operator, code="S-NODE")

    created = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"]),
        operator,
    )

    assert created["code"] == "N001"
    assert len(created["material_inputs"]) == 1
    assert len(created["consumables"]) == 1
    assert len(created["public_services"]) == 1
    assert len(created["equipment"]) == 1
    assert len(created["outputs"]) == 1
    assert Decimal(created["outputs"][0]["output_per_ton"]) == Decimal("0.8")

    page = service.list_nodes(keyword="Leaching", node_type="hydrometallurgy", status="enabled", page=1, page_size=10)
    assert page["total"] == 1
    assert page["items"][0]["id"] == created["id"]

    updated = service.update_node(
        created["id"],
        ProcessNodeUpdateWithChildren(
            name="Updated Leaching",
            status="enabled",
            material_inputs=[
                {"material_id": material["id"], "amount_per_ton": Decimal("2.5"), "unit": "t", "sort_order": 1},
            ],
            consumables=[],
            public_services=[],
            equipment=[],
            outputs=[
                {"product_id": product["id"], "output_per_ton": Decimal("0.9"), "unit": "kg", "is_main_product": True, "sort_order": 1},
            ],
        ),
        operator,
    )

    assert updated["name"] == "Updated Leaching"
    assert len(updated["material_inputs"]) == 1
    assert Decimal(updated["material_inputs"][0]["amount_per_ton"]) == Decimal("2.5")
    assert updated["consumables"] == []
    assert updated["public_services"] == []
    assert updated["equipment"] == []
    assert Decimal(updated["outputs"][0]["output_per_ton"]) == Decimal("0.9")

    service.delete_node(created["id"], operator)
    with pytest.raises(AppException, match="数据不存在"):
        service.get_node(created["id"])


def test_node_enabled_validation_requires_outputs_and_enabled_references(db_session: Session) -> None:
    """Enabled nodes require output products and enabled referenced libraries."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    material = create_material(service, operator, code="M-VAL")
    product = create_product(service, operator, code="P-VAL")
    consumable = create_consumable(service, operator, code="C-VAL")
    public_service = create_public_service(service, operator, code="S-VAL")

    payload = node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-NO-OUTPUT")
    payload.outputs = []
    with pytest.raises(AppException, match="启用节点时至少需要配置一个输出产品"):
        service.create_node(payload, operator)

    service.update_status("product", product["id"], ProcessLibraryStatusUpdate(status="disabled"), operator)
    with pytest.raises(AppException, match="引用的产品必须为启用状态"):
        service.create_node(
            node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-DISABLED-PRODUCT"),
            operator,
        )

    with pytest.raises(AppException, match="引用的原料不存在或已删除"):
        service.create_node(
            node_payload(9999, product["id"], consumable["id"], public_service["id"], code="N-MISSING-MATERIAL", status="draft"),
            operator,
        )


def test_node_delete_blocks_route_references(db_session: Session) -> None:
    """A node referenced by an active route cannot be deleted."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    material = create_material(service, operator, code="M-ROUTE-NODE")
    product = create_product(service, operator, code="P-ROUTE-NODE")
    consumable = create_consumable(service, operator, code="C-ROUTE-NODE")
    public_service = create_public_service(service, operator, code="S-ROUTE-NODE")
    node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-REF"),
        operator,
    )

    route = ProcessRoute(
        code="R-NODE-REF",
        name="Route",
        input_material_id=material["id"],
        final_product_id=product["id"],
        version="1.0",
        status="enabled",
        sort_order=1,
        is_deleted=False,
    )
    db_session.add(route)
    db_session.flush()
    db_session.add(ProcessRouteNode(route_id=route.id, node_id=node["id"], sort_order=1, is_deleted=False))
    db_session.commit()

    with pytest.raises(AppException, match="当前节点已被工艺路线引用，不能删除"):
        service.delete_node(node["id"], operator)


def test_route_crud_snapshot_copy_and_delete(db_session: Session) -> None:
    """Routes persist node chains, snapshots, copies, and soft deletes correctly."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    material = create_material(service, operator, code="M-ROUTE")
    product = create_product(service, operator, code="P-ROUTE")
    consumable = create_consumable(service, operator, code="C-ROUTE")
    public_service = create_public_service(service, operator, code="S-ROUTE")
    first_node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-1", name="Route Node 1"),
        operator,
    )
    second_node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-2", name="Route Node 2"),
        operator,
    )

    created = service.create_route(route_payload(material["id"], product["id"], [first_node["id"], second_node["id"]]), operator)

    assert created["route"]["code"] == "R001"
    assert len(created["nodes"]) == 2
    assert created["nodes"][0]["node"]["code"] == "N-ROUTE-1"

    page = service.list_routes(keyword="Route", status="enabled", page=1, page_size=10)
    assert page["total"] == 1
    assert page["items"][0]["node_count"] == 2

    reordered = service.reorder_route_nodes(
        created["route"]["id"],
        ProcessRouteNodeReorderPayload(
            items=[
                {"route_node_id": created["nodes"][0]["id"], "sort_order": 2},
                {"route_node_id": created["nodes"][1]["id"], "sort_order": 1},
            ]
        ),
        operator,
    )
    assert reordered[0]["node_id"] == second_node["id"]

    updated = service.update_route(
        created["route"]["id"],
        ProcessRouteUpdateWithNodes(
            name="Updated Route",
            nodes=[
                {"node_id": second_node["id"], "sort_order": 1, "node_params_json": {"recovery": "93%"}, "remark": "step 2"},
                {"node_id": first_node["id"], "sort_order": 2, "node_params_json": {"recovery": "91%"}, "remark": "step 1"},
            ],
        ),
        operator,
    )
    assert updated["route"]["name"] == "Updated Route"
    assert updated["nodes"][0]["node"]["code"] == "N-ROUTE-2"
    assert json.loads(updated["nodes"][0]["node_params_json"])["recovery"] == "93%"

    version = service.create_route_version(
        created["route"]["id"],
        ProcessRouteVersionCreatePayload(change_log="initial snapshot"),
        operator,
    )
    assert version["version_no"] == 1
    assert json.loads(version["snapshot_json"])["route"]["name"] == "Updated Route"

    versions = service.list_route_versions(created["route"]["id"])
    assert versions[0]["version_no"] == 1

    copied = service.copy_route(created["route"]["id"], operator)
    assert copied["route"]["status"] == "draft"
    assert copied["route"]["code"].startswith("R001_COPY")
    assert len(copied["nodes"]) == 2

    service.delete_route_node(created["route"]["id"], updated["nodes"][1]["id"], operator)
    remaining = service.get_route(created["route"]["id"])
    assert len(remaining["nodes"]) == 1

    service.delete_route(created["route"]["id"], operator)
    with pytest.raises(AppException, match="数据不存在"):
        service.get_route(created["route"]["id"])


def test_route_enabled_validation_and_sort_order_rules(db_session: Session) -> None:
    """Enabled routes require enabled references, at least one node, and unique sort order."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    material = create_material(service, operator, code="M-ROUTE-VAL")
    product = create_product(service, operator, code="P-ROUTE-VAL")
    consumable = create_consumable(service, operator, code="C-ROUTE-VAL")
    public_service = create_public_service(service, operator, code="S-ROUTE-VAL")
    enabled_node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-VAL-1"),
        operator,
    )

    with pytest.raises(AppException, match="启用路线时至少需要配置一个节点"):
        service.create_route(
            ProcessRouteCreateWithNodes(
                code="R-NO-NODE",
                name="No Node Route",
                input_material_id=material["id"],
                final_product_id=product["id"],
                status="enabled",
                nodes=[],
            ),
            operator,
        )

    disabled_node = service.create_node(
        node_payload(
            material["id"],
            product["id"],
            consumable["id"],
            public_service["id"],
            code="N-ROUTE-VAL-2",
            status="draft",
        ),
        operator,
    )
    with pytest.raises(AppException, match="启用路线时，引用的工艺节点必须为启用状态"):
        service.create_route(route_payload(material["id"], product["id"], [disabled_node["id"]], code="R-DISABLED-NODE"), operator)

    service.update_status("material", material["id"], ProcessLibraryStatusUpdate(status="disabled"), operator)
    with pytest.raises(AppException, match="启用路线时，引用的输入原料必须为启用状态"):
        service.create_route(route_payload(material["id"], product["id"], [enabled_node["id"]], code="R-DISABLED-MATERIAL"), operator)
    service.update_status("material", material["id"], ProcessLibraryStatusUpdate(status="enabled"), operator)

    created = service.create_route(route_payload(material["id"], product["id"], [enabled_node["id"]], code="R-VAL-OK"), operator)

    with pytest.raises(AppException, match="启用路线时至少需要配置一个节点"):
        service.delete_route_node(created["route"]["id"], created["nodes"][0]["id"], operator)

    third_node = service.create_node(
        node_payload(material["id"], product["id"], consumable["id"], public_service["id"], code="N-ROUTE-VAL-3"),
        operator,
    )
    created = service.update_route(
        created["route"]["id"],
        ProcessRouteUpdateWithNodes(
            nodes=[
                {"node_id": enabled_node["id"], "sort_order": 1},
                {"node_id": third_node["id"], "sort_order": 2},
            ]
        ),
        operator,
    )
    with pytest.raises(AppException, match="节点顺序不能重复"):
        service.reorder_route_nodes(
            created["route"]["id"],
            ProcessRouteNodeReorderPayload(
                items=[
                    {"route_node_id": created["nodes"][0]["id"], "sort_order": 1},
                    {"route_node_id": created["nodes"][1]["id"], "sort_order": 1},
                ]
            ),
            operator,
        )


def test_code_unique_includes_soft_deleted_records(db_session: Session) -> None:
    """Soft-deleted records still reserve their code."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)

    created = create_product(service, operator, code="P-UNIQUE")
    service.delete_library("product", created["id"], operator)

    with pytest.raises(AppException, match="编码已存在"):
        create_product(service, operator, code="P-UNIQUE")


def test_delete_blocks_node_and_route_references(db_session: Session) -> None:
    """Referenced materials, products, consumables and services cannot be deleted."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    material = create_material(service, operator, code="M-REF")
    product = create_product(service, operator, code="P-REF")
    consumable = create_consumable(service, operator, code="C-REF")
    public_service = create_public_service(service, operator, code="S-REF")

    node = ProcessNode(
        code="N-REF",
        name="Leaching",
        node_type="hydrometallurgy",
        staff=Decimal("1"),
        area=Decimal("20"),
        status="enabled",
        version="1.0",
        sort_order=1,
        is_deleted=False,
    )
    db_session.add(node)
    db_session.flush()
    db_session.add_all(
        [
            ProcessNodeMaterialInput(node_id=node.id, material_id=material["id"], amount_per_ton=Decimal("1"), unit="kg", sort_order=1, is_deleted=False),
            ProcessNodeOutput(node_id=node.id, product_id=product["id"], output_per_ton=Decimal("1"), unit="kg", is_main_product=True, sort_order=1, is_deleted=False),
            ProcessNodeConsumable(node_id=node.id, consumable_id=consumable["id"], amount_per_ton=Decimal("1"), unit="kg", sort_order=1, is_deleted=False),
            ProcessNodePublicService(node_id=node.id, public_service_id=public_service["id"], amount_per_ton=Decimal("1"), unit="kWh", sort_order=1, is_deleted=False),
            ProcessRoute(
                code="R-REF",
                name="Route",
                input_material_id=material["id"],
                final_product_id=product["id"],
                version="1.0",
                status="enabled",
                sort_order=1,
                is_deleted=False,
            ),
        ]
    )
    db_session.commit()

    for kind, item_id in [
        ("material", material["id"]),
        ("product", product["id"]),
        ("consumable", consumable["id"]),
        ("public_service", public_service["id"]),
    ]:
        with pytest.raises(AppException, match="当前数据已被引用，不能删除"):
            service.delete_library(kind, item_id, operator)
