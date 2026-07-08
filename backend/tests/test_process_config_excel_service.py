"""Process configuration Excel import/export tests."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
import sys
from pathlib import Path

from openpyxl import load_workbook
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR))

from app.api.deps import get_current_user  # noqa: E402
from app.core.database import get_db  # noqa: E402
from app.models import Base, Role, User  # noqa: E402
from app.schemas.process_config import (  # noqa: E402
    ProcessConsumableCreateWithPrices,
    ProcessLibraryRegionPricePayload,
    ProcessMaterialCreateWithPrices,
    ProcessProductCreateWithPrices,
    ProcessPublicServiceCreateWithPrices,
)
from app.services.process_config_excel_service import ProcessConfigExcelService  # noqa: E402
from app.services.process_config_service import ProcessConfigService  # noqa: E402
from main import app  # noqa: E402


@pytest.fixture()
def db_session() -> Session:
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    with session_factory() as session:
        yield session
    engine.dispose()


@pytest.fixture()
def api_db_session() -> Session:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    session = session_factory()
    try:
        yield session
    finally:
        app.dependency_overrides.clear()
        session.close()
        engine.dispose()


def seed_operator(db: Session) -> User:
    operator = User(username="admin", password_hash="x", real_name="Admin", status="enabled")
    db.add(operator)
    db.commit()
    db.refresh(operator)
    return operator


def region_prices(asia_price: str = "10.5") -> list[ProcessLibraryRegionPricePayload]:
    return [
        ProcessLibraryRegionPricePayload(region_code="asia", unit_price=Decimal(asia_price), unit="kg"),
        ProcessLibraryRegionPricePayload(region_code="europe", unit_price=Decimal("20.5"), unit="kg"),
        ProcessLibraryRegionPricePayload(region_code="americas", unit_price=Decimal("30.5"), unit="kg"),
    ]


def workbook_bytes(workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_material(service: ProcessConfigService, operator: User, code: str) -> dict:
    return service.create_library(
        "material",
        ProcessMaterialCreateWithPrices(
            code=code,
            name=f"Material {code}",
            type="battery",
            description="Cathode feed",
            unit="kg",
            status="enabled",
            region_prices=region_prices(),
        ),
        operator,
    )


def create_product(service: ProcessConfigService, operator: User, code: str) -> dict:
    return service.create_library(
        "product",
        ProcessProductCreateWithPrices(
            code=code,
            name=f"Product {code}",
            type="metal",
            description="Final product",
            unit="kg",
            status="enabled",
        ),
        operator,
    )


def create_consumable(service: ProcessConfigService, operator: User, code: str) -> dict:
    return service.create_library(
        "consumable",
        ProcessConsumableCreateWithPrices(
            code=code,
            name=f"Consumable {code}",
            type="chemical",
            description="Acid",
            unit="kg",
            status="enabled",
        ),
        operator,
    )


def create_public_service(service: ProcessConfigService, operator: User, code: str) -> dict:
    return service.create_library(
        "public_service",
        ProcessPublicServiceCreateWithPrices(
            code=code,
            name=f"Service {code}",
            type="utility",
            description="Power",
            unit="kWh",
            status="enabled",
        ),
        operator,
    )


def test_material_excel_api_roundtrip(api_db_session: Session) -> None:
    """Admin can download the material template, import Excel data, and export current records."""

    operator = seed_operator(api_db_session)
    admin_role = Role(name="Admin", code="admin", enabled=True)
    operator.roles = [admin_role]
    api_db_session.add(admin_role)
    api_db_session.commit()

    def override_db() -> Session:
        return api_db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: operator
    client = TestClient(app)

    template_response = client.get("/api/process-config/materials/template")
    assert template_response.status_code == 200
    assert template_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(template_response.content))
    workbook["原料库"].append(
        ["M-EXCEL-API", "Excel Material", "battery", "From API import", "kg", "enabled", 1, "remark", "1.1", "2.2", "3.3"]
    )

    import_response = client.post(
        "/api/process-config/materials/import",
        files={
            "file": (
                "materials.xlsx",
                workbook_bytes(workbook),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    import_data = import_response.json()["data"]
    assert import_data["module"] == "materials"
    assert import_data["imported_count"] == 1
    assert import_data["imported_codes"] == ["M-EXCEL-API"]

    export_response = client.get("/api/process-config/materials/export", params={"keyword": "M-EXCEL-API"})
    assert export_response.status_code == 200
    export_workbook = load_workbook(BytesIO(export_response.content), data_only=True)
    rows = list(export_workbook["原料库"].iter_rows(values_only=True))
    assert rows[1][0] == "M-EXCEL-API"
    assert rows[1][1] == "Excel Material"


def test_material_excel_service_template_import_and_export_roundtrip(db_session: Session) -> None:
    """Material Excel service supports template generation, import validation, and export."""

    operator = seed_operator(db_session)
    excel_service = ProcessConfigExcelService(db_session)
    service = ProcessConfigService(db_session)

    template_bytes, template_name = excel_service.build_template("materials")
    assert template_name == "process-materials-template.xlsx"

    workbook = load_workbook(BytesIO(template_bytes))
    assert workbook.sheetnames[:2] == ["原料库", "填写说明"]
    workbook["原料库"].append(
        ["M-EXCEL-SVC", "Excel Service Material", "battery", "From service import", "kg", "enabled", 2, "remark", "11.1", "22.2", "33.3"]
    )

    result = excel_service.import_module("materials", workbook_bytes(workbook), operator)
    assert result["module"] == "materials"
    assert result["imported_count"] == 1
    assert result["imported_codes"] == ["M-EXCEL-SVC"]

    page = service.list_library("material", keyword="M-EXCEL-SVC", page=1, page_size=10)
    assert page["total"] == 1
    assert page["items"][0]["name"] == "Excel Service Material"

    export_bytes, export_name = excel_service.export_module("materials", {"keyword": "M-EXCEL-SVC"})
    assert export_name.startswith("process-materials-export-")
    export_workbook = load_workbook(BytesIO(export_bytes), data_only=True)
    export_rows = list(export_workbook["原料库"].iter_rows(values_only=True))
    assert export_rows[1][0] == "M-EXCEL-SVC"
    assert Decimal(str(export_rows[1][8])) == Decimal("11.1")


def test_node_and_route_excel_service_import_roundtrip(db_session: Session) -> None:
    """Node and route Excel imports persist child sheets and node chains correctly."""

    operator = seed_operator(db_session)
    service = ProcessConfigService(db_session)
    excel_service = ProcessConfigExcelService(db_session)

    material = create_material(service, operator, code="M-EXCEL-NODE")
    product = create_product(service, operator, code="P-EXCEL-NODE")
    consumable = create_consumable(service, operator, code="C-EXCEL-NODE")
    public_service = create_public_service(service, operator, code="S-EXCEL-NODE")

    node_template_bytes, _ = excel_service.build_template("nodes")
    node_workbook = load_workbook(BytesIO(node_template_bytes))
    node_workbook["节点基础信息"].append(["N-EXCEL-1", "Excel Node", "hydrometallurgy", "2", "120", "Excel node", "enabled", "v1.0", 1, "remark"])
    node_workbook["节点输入原料"].append(["N-EXCEL-1", material["code"], "1.5", "t", 1, ""])
    node_workbook["节点消耗品"].append(["N-EXCEL-1", consumable["code"], "0.2", "kg", 1, ""])
    node_workbook["节点公共服务"].append(["N-EXCEL-1", public_service["code"], "10", "kWh", 1, ""])
    node_workbook["节点设备投资"].append(["N-EXCEL-1", "Reactor", "tank", "2", "100000", "CNY", 1, "main"])
    node_workbook["节点输出产品"].append(["N-EXCEL-1", product["code"], "0.8", "kg", True, 1, ""])

    node_result = excel_service.import_module("nodes", workbook_bytes(node_workbook), operator)
    assert node_result["module"] == "nodes"
    assert node_result["imported_count"] == 1
    assert node_result["imported_codes"] == ["N-EXCEL-1"]

    node_page = service.list_nodes(keyword="N-EXCEL-1", page=1, page_size=10)
    assert node_page["total"] == 1
    node_detail = service.get_node(node_page["items"][0]["id"])
    assert len(node_detail["material_inputs"]) == 1
    assert len(node_detail["outputs"]) == 1

    route_template_bytes, _ = excel_service.build_template("routes")
    route_workbook = load_workbook(BytesIO(route_template_bytes))
    route_workbook["路线基础信息"].append(["R-EXCEL-1", "Excel Route", material["code"], product["code"], "v1.0", "Excel route", "enabled", 1, "remark"])
    route_workbook["路线节点链路"].append(["R-EXCEL-1", "N-EXCEL-1", 1, '{"recovery": 95}', "step 1"])

    route_result = excel_service.import_module("routes", workbook_bytes(route_workbook), operator)
    assert route_result["module"] == "routes"
    assert route_result["imported_count"] == 1
    assert route_result["imported_codes"] == ["R-EXCEL-1"]

    route_page = service.list_routes(keyword="R-EXCEL-1", page=1, page_size=10)
    assert route_page["total"] == 1
    route_detail = service.get_route(route_page["items"][0]["id"])
    assert route_detail["route"]["code"] == "R-EXCEL-1"
    assert len(route_detail["nodes"]) == 1
    assert route_detail["nodes"][0]["node"]["code"] == "N-EXCEL-1"
