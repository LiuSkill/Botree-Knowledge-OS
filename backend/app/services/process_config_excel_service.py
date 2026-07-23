"""Process configuration Excel import/export services."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import json
import logging
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.exceptions import AppException
from app.models.user import User
from app.repositories.process_config_repository import ProcessNodeRepository, ProcessRouteRepository
from app.schemas.process_config import (
    ProcessConfigImportResultOut,
    ProcessLibraryCreateWithPrices,
    ProcessLibraryRegionPricePayload,
    ProcessNodeConsumablePayload,
    ProcessNodeCreateWithChildren,
    ProcessNodeEquipmentPayload,
    ProcessNodeMaterialInputPayload,
    ProcessNodeOutputPayload,
    ProcessNodePublicServicePayload,
    ProcessRouteCreateWithNodes,
    ProcessRouteNodePayload,
)
from app.services.process_config_service import ProcessConfigService, ProcessLibraryKind

logger = logging.getLogger(__name__)

ProcessExcelModule = Literal["materials", "products", "consumables", "public-services", "nodes", "routes"]

STATUS_LABELS = {"enabled": "启用", "draft": "草稿", "disabled": "停用"}
STATUS_ALIASES = {
    "enabled": "enabled",
    "启用": "enabled",
    "draft": "draft",
    "草稿": "draft",
    "disabled": "disabled",
    "停用": "disabled",
}
NODE_TYPE_LABELS = {
    "pretreatment": "预处理",
    "hydrometallurgy": "湿法冶金",
    "pyrometallurgy": "火法冶金",
    "post_treatment": "后处理",
}
NODE_TYPE_ALIASES = {
    "pretreatment": "pretreatment",
    "预处理": "pretreatment",
    "hydrometallurgy": "hydrometallurgy",
    "湿法冶金": "hydrometallurgy",
    "pyrometallurgy": "pyrometallurgy",
    "火法冶金": "pyrometallurgy",
    "post_treatment": "post_treatment",
    "后处理": "post_treatment",
}
BOOL_ALIASES = {
    "1": True,
    "true": True,
    "yes": True,
    "y": True,
    "是": True,
    "0": False,
    "false": False,
    "no": False,
    "n": False,
    "否": False,
}
REGION_COLUMNS = (
    ("asia", "亚洲单价(CNY)"),
    ("europe", "欧洲单价(EUR)"),
    ("americas", "美洲单价(USD)"),
)
LIBRARY_HEADERS = ["编码", "名称", "类型", "描述", "单位", "状态", "排序", "备注", *(column for _, column in REGION_COLUMNS)]
NODE_SHEET_HEADERS: dict[str, list[str]] = {
    "节点基础信息": ["节点编码", "节点名称", "节点类型", "人员", "占地面积", "描述", "状态", "版本", "排序", "备注"],
    "节点输入原料": ["节点编码", "原料编码", "投入量/吨", "单位", "排序", "备注"],
    "节点消耗品": ["节点编码", "消耗品编码", "消耗量/吨", "单位", "排序", "备注"],
    "节点公共服务": ["节点编码", "公共服务编码", "消耗量/吨", "单位", "排序", "备注"],
    "节点设备投资": ["节点编码", "设备名称", "设备类型", "数量", "投资金额", "币种", "排序", "备注"],
    "节点输出产品": ["节点编码", "产品编码", "产出量/吨", "单位", "是否主产品", "排序", "备注"],
}
ROUTE_SHEET_HEADERS: dict[str, list[str]] = {
    "路线基础信息": ["路线编码", "路线名称", "输入原料编码", "最终产品编码", "版本", "描述", "状态", "排序", "备注"],
    "路线节点链路": ["路线编码", "节点编码", "排序", "节点参数JSON", "备注"],
}
MAX_EXPORT_PAGE_SIZE = 1000000


@dataclass(frozen=True)
class ProcessExcelModuleConfig:
    """Excel 模块配置。"""

    module: ProcessExcelModule
    label: str
    filename_prefix: str
    library_kind: ProcessLibraryKind | None = None


@dataclass
class ParsedNodeRow:
    """节点基础信息解析结果。"""

    row_no: int
    code: str
    data: dict[str, Any]
    material_inputs: list[ProcessNodeMaterialInputPayload]
    consumables: list[ProcessNodeConsumablePayload]
    public_services: list[ProcessNodePublicServicePayload]
    equipment: list[ProcessNodeEquipmentPayload]
    outputs: list[ProcessNodeOutputPayload]


@dataclass
class ParsedRouteRow:
    """路线基础信息解析结果。"""

    row_no: int
    code: str
    data: dict[str, Any]
    nodes: list[ProcessRouteNodePayload]


MODULE_CONFIGS: dict[ProcessExcelModule, ProcessExcelModuleConfig] = {
    "materials": ProcessExcelModuleConfig("materials", "原料库", "process-materials", "material"),
    "products": ProcessExcelModuleConfig("products", "产品库", "process-products", "product"),
    "consumables": ProcessExcelModuleConfig("consumables", "消耗品库", "process-consumables", "consumable"),
    "public-services": ProcessExcelModuleConfig("public-services", "公共服务库", "process-public-services", "public_service"),
    "nodes": ProcessExcelModuleConfig("nodes", "工艺节点库", "process-nodes"),
    "routes": ProcessExcelModuleConfig("routes", "工艺路线库", "process-routes"),
}


class ProcessConfigExcelService:
    """工艺配置 Excel 模板、导出与导入服务。"""

    def __init__(self, db: Session) -> None:
        self.db = db
        self.process_service = ProcessConfigService(db)

    def build_template(self, module: ProcessExcelModule) -> tuple[bytes, str]:
        """生成模块导入模板。"""

        config = self._module_config(module)
        workbook = self._build_template_workbook(config)
        logger.info("工艺配置模板生成完成: module=%s", module)
        return self._workbook_bytes(workbook), f"{config.filename_prefix}-template.xlsx"

    def export_module(self, module: ProcessExcelModule, filters: dict[str, Any] | None = None) -> tuple[bytes, str]:
        """导出模块当前数据。"""

        config = self._module_config(module)
        workbook = self._build_export_workbook(config, filters or {})
        file_name = f"{config.filename_prefix}-export-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        logger.info("工艺配置数据导出完成: module=%s", module)
        return self._workbook_bytes(workbook), file_name

    def import_module(self, module: ProcessExcelModule, content: bytes, operator: User) -> dict[str, Any]:
        """导入模块数据，默认整批事务提交。"""

        config = self._module_config(module)
        workbook = self._load_workbook(content)
        if config.library_kind:
            result = self._import_library(config, workbook, operator)
        elif module == "nodes":
            result = self._import_nodes(config, workbook, operator)
        else:
            result = self._import_routes(config, workbook, operator)
        logger.info("工艺配置 Excel 导入完成: module=%s imported_count=%s operator_id=%s", module, result["imported_count"], operator.id)
        return result

    def _build_template_workbook(self, config: ProcessExcelModuleConfig) -> Workbook:
        if config.library_kind:
            workbook = self._create_workbook(config.label, LIBRARY_HEADERS)
            self._append_instruction_sheet(
                workbook,
                [
                    ("编码", "全局唯一，若数据库已存在相同编码（含软删除数据）将报错。"),
                    ("状态", "支持填写：启用 / 草稿 / 停用，或 enabled / draft / disabled。"),
                    ("区域单价", "亚洲/欧洲/美洲分别对应 CNY / EUR / USD。"),
                ],
            )
            return workbook

        workbook = Workbook()
        workbook.remove(workbook.active)
        if config.module == "nodes":
            for sheet_name, headers in NODE_SHEET_HEADERS.items():
                self._create_sheet(workbook, sheet_name, headers)
            self._append_instruction_sheet(
                workbook,
                [
                    ("节点类型", "支持填写：预处理/湿法冶金/火法冶金/后处理，或对应英文编码。"),
                    ("状态", "支持填写：启用 / 草稿 / 停用，或 enabled / draft / disabled。"),
                    ("子表关联", "除“节点基础信息”外，其他 Sheet 通过“节点编码”关联。"),
                    ("引用字段", "原料/产品/消耗品/公共服务请填写各自库中的“编码”。"),
                ],
            )
            return workbook

        for sheet_name, headers in ROUTE_SHEET_HEADERS.items():
            self._create_sheet(workbook, sheet_name, headers)
        self._append_instruction_sheet(
            workbook,
            [
                ("状态", "支持填写：启用 / 草稿 / 停用，或 enabled / draft / disabled。"),
                ("子表关联", "“路线节点链路”通过“路线编码”关联“路线基础信息”。"),
                ("节点参数JSON", "留空或填写合法 JSON，例如 {\"recovery\": 90}。"),
                ("引用字段", "输入原料、最终产品、节点请填写各自库中的“编码”。"),
            ],
        )
        return workbook

    def _build_export_workbook(self, config: ProcessExcelModuleConfig, filters: dict[str, Any]) -> Workbook:
        if config.library_kind:
            return self._export_library_workbook(config, filters)
        if config.module == "nodes":
            return self._export_nodes_workbook(filters)
        return self._export_routes_workbook(filters)

    def _export_library_workbook(self, config: ProcessExcelModuleConfig, filters: dict[str, Any]) -> Workbook:
        result = self.process_service.list_library(
            config.library_kind,
            keyword=self._optional_text(filters.get("keyword")),
            type_code=self._optional_text(filters.get("type")),
            output_type=self._optional_text(filters.get("output_type")),
            status=self._optional_text(filters.get("status")),
            page=1,
            page_size=MAX_EXPORT_PAGE_SIZE,
        )
        workbook = self._create_workbook(config.label, LIBRARY_HEADERS)
        sheet = workbook.active
        for item in result["items"]:
            prices = {price["region_code"]: price for price in item.get("region_prices", [])}
            sheet.append(
                [
                    item["code"],
                    item["name"],
                    item["type"],
                    item.get("description") or "",
                    item["unit"],
                    self._status_label(item["status"]),
                    item["sort_order"],
                    item.get("remark") or "",
                    prices.get("asia", {}).get("unit_price", 0),
                    prices.get("europe", {}).get("unit_price", 0),
                    prices.get("americas", {}).get("unit_price", 0),
                ]
            )
        self._finalize_sheet(sheet)
        return workbook

    def _export_nodes_workbook(self, filters: dict[str, Any]) -> Workbook:
        result = self.process_service.list_nodes(
            keyword=self._optional_text(filters.get("keyword")),
            node_type=self._optional_text(filters.get("node_type")),
            status=self._optional_text(filters.get("status")),
            page=1,
            page_size=MAX_EXPORT_PAGE_SIZE,
        )
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = {name: self._create_sheet(workbook, name, headers) for name, headers in NODE_SHEET_HEADERS.items()}
        material_repo = self.process_service._repository("material")
        product_repo = self.process_service._repository("product")
        consumable_repo = self.process_service._repository("consumable")
        public_service_repo = self.process_service._repository("public_service")

        for item in result["items"]:
            detail = self.process_service.get_node(item["id"])
            sheets["节点基础信息"].append(
                [
                    detail["code"],
                    detail["name"],
                    self._node_type_label(detail["node_type"]),
                    detail["staff"],
                    detail["area"],
                    detail.get("description") or "",
                    self._status_label(detail["status"]),
                    detail["version"],
                    detail["sort_order"],
                    detail.get("remark") or "",
                ]
            )
            for child in detail["material_inputs"]:
                material = material_repo.get_by_id(child["material_id"])
                sheets["节点输入原料"].append(
                    [
                        detail["code"],
                        material.code if material else "",
                        child["amount_per_ton"],
                        child["unit"],
                        child["sort_order"],
                        child.get("remark") or "",
                    ]
                )
            for child in detail["consumables"]:
                consumable = consumable_repo.get_by_id(child["consumable_id"])
                sheets["节点消耗品"].append(
                    [
                        detail["code"],
                        consumable.code if consumable else "",
                        child["amount_per_ton"],
                        child["unit"],
                        child["sort_order"],
                        child.get("remark") or "",
                    ]
                )
            for child in detail["public_services"]:
                public_service = public_service_repo.get_by_id(child["public_service_id"])
                sheets["节点公共服务"].append(
                    [
                        detail["code"],
                        public_service.code if public_service else "",
                        child["amount_per_ton"],
                        child["unit"],
                        child["sort_order"],
                        child.get("remark") or "",
                    ]
                )
            for child in detail["equipment"]:
                sheets["节点设备投资"].append(
                    [
                        detail["code"],
                        child["equipment_name"],
                        child.get("equipment_type") or "",
                        child["quantity"],
                        child["sort_order"],
                        child.get("remark") or "",
                    ]
                )
            for child in detail["outputs"]:
                product = product_repo.get_by_id(child["product_id"])
                sheets["节点输出产品"].append(
                    [
                        detail["code"],
                        product.code if product else "",
                        child["output_per_ton"],
                        child["unit"],
                        "是" if child["is_main_product"] else "否",
                        child["sort_order"],
                        child.get("remark") or "",
                    ]
                )

        for sheet in sheets.values():
            self._finalize_sheet(sheet)
        return workbook

    def _export_routes_workbook(self, filters: dict[str, Any]) -> Workbook:
        result = self.process_service.list_routes(
            keyword=self._optional_text(filters.get("keyword")),
            status=self._optional_text(filters.get("status")),
            input_material_id=self._optional_int(filters.get("input_material_id")),
            final_product_id=self._optional_int(filters.get("final_product_id")),
            page=1,
            page_size=MAX_EXPORT_PAGE_SIZE,
        )
        workbook = Workbook()
        workbook.remove(workbook.active)
        base_sheet = self._create_sheet(workbook, "路线基础信息", ROUTE_SHEET_HEADERS["路线基础信息"])
        node_sheet = self._create_sheet(workbook, "路线节点链路", ROUTE_SHEET_HEADERS["路线节点链路"])

        for item in result["items"]:
            detail = self.process_service.get_route(item["id"])
            route = detail["route"]
            base_sheet.append(
                [
                    route["code"],
                    route["name"],
                    detail["input_material"]["code"],
                    detail["final_product"]["code"],
                    route["version"],
                    route.get("description") or "",
                    self._status_label(route["status"]),
                    route["sort_order"],
                    route.get("remark") or "",
                ]
            )
            for route_node in detail["nodes"]:
                node_sheet.append(
                    [
                        route["code"],
                        route_node["node"]["code"],
                        route_node["sort_order"],
                        route_node.get("node_params_json") or "",
                        route_node.get("remark") or "",
                    ]
                )

        self._finalize_sheet(base_sheet)
        self._finalize_sheet(node_sheet)
        return workbook

    def _import_library(self, config: ProcessExcelModuleConfig, workbook: Workbook, operator: User) -> dict[str, Any]:
        rows = self._sheet_rows(workbook, config.label, LIBRARY_HEADERS)
        repo = self.process_service._repository(config.library_kind)
        errors: list[dict[str, Any]] = []
        seen_codes: set[str] = set()
        payloads: list[ProcessLibraryCreateWithPrices] = []

        for row_no, row in rows:
            code = self._required_text(row, "编码", config.label, row_no, errors)
            name = self._required_text(row, "名称", config.label, row_no, errors)
            item_type = self._required_text(row, "类型", config.label, row_no, errors)
            unit = self._required_text(row, "单位", config.label, row_no, errors)
            status = self._parse_status(row.get("状态"), config.label, row_no, "状态", errors, default="enabled")
            sort_order = self._parse_int(row.get("排序"), config.label, row_no, "排序", errors, default=0)
            if code:
                if code in seen_codes:
                    self._add_error(errors, config.label, row_no, "编码", "同一导入文件中编码重复")
                else:
                    seen_codes.add(code)
                if repo.get_by_code(code):
                    self._add_error(errors, config.label, row_no, "编码", "编码已存在")
            if not all(value is not None for value in (code, name, item_type, unit, status, sort_order)):
                continue

            region_prices = []
            for region_code, column in REGION_COLUMNS:
                price = self._parse_decimal(row.get(column), config.label, row_no, column, errors, default=Decimal("0"))
                if price is None:
                    continue
                region_prices.append(
                    ProcessLibraryRegionPricePayload(
                        region_code=region_code,
                        unit_price=price,
                        unit=unit,
                        status=status,
                    )
                )

            if len(region_prices) != len(REGION_COLUMNS):
                continue

            try:
                payloads.append(
                    ProcessLibraryCreateWithPrices(
                        code=code,
                        name=name,
                        type=item_type,
                        description=self._optional_text(row.get("描述")),
                        unit=unit,
                        status=status,
                        sort_order=sort_order,
                        remark=self._optional_text(row.get("备注")),
                        region_prices=region_prices,
                    )
                )
            except ValidationError as exc:
                self._append_validation_errors(errors, config.label, row_no, exc)

        if errors:
            self._raise_import_errors(errors)

        imported_codes: list[str] = []
        try:
            for payload in payloads:
                self._persist_library(config.library_kind, payload, operator)
                imported_codes.append(payload.code)
            self.process_service.system_service.record_operation(
                operator,
                f"导入{config.label}",
                f"process_{config.library_kind}",
                detail=f"批量导入{config.label} {len(imported_codes)} 条",
            )
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise
        return ProcessConfigImportResultOut(module=config.module, imported_count=len(imported_codes), imported_codes=imported_codes).model_dump()

    def _import_nodes(self, config: ProcessExcelModuleConfig, workbook: Workbook, operator: User) -> dict[str, Any]:
        errors: list[dict[str, Any]] = []
        base_rows = self._sheet_rows(workbook, "节点基础信息", NODE_SHEET_HEADERS["节点基础信息"])
        material_rows = self._sheet_rows(workbook, "节点输入原料", NODE_SHEET_HEADERS["节点输入原料"])
        consumable_rows = self._sheet_rows(workbook, "节点消耗品", NODE_SHEET_HEADERS["节点消耗品"])
        public_service_rows = self._sheet_rows(workbook, "节点公共服务", NODE_SHEET_HEADERS["节点公共服务"])
        equipment_rows = self._sheet_rows(workbook, "节点设备投资", NODE_SHEET_HEADERS["节点设备投资"])
        output_rows = self._sheet_rows(workbook, "节点输出产品", NODE_SHEET_HEADERS["节点输出产品"])

        node_repo = ProcessNodeRepository(self.db)
        material_repo = self.process_service._repository("material")
        product_repo = self.process_service._repository("product")
        consumable_repo = self.process_service._repository("consumable")
        public_service_repo = self.process_service._repository("public_service")

        parsed_nodes: dict[str, ParsedNodeRow] = {}
        seen_codes: set[str] = set()

        for row_no, row in base_rows:
            code = self._required_text(row, "节点编码", "节点基础信息", row_no, errors)
            name = self._required_text(row, "节点名称", "节点基础信息", row_no, errors)
            node_type = self._parse_node_type(row.get("节点类型"), "节点基础信息", row_no, "节点类型", errors)
            staff = self._parse_decimal(row.get("人员"), "节点基础信息", row_no, "人员", errors, default=Decimal("0"))
            area = self._parse_decimal(row.get("占地面积"), "节点基础信息", row_no, "占地面积", errors, default=Decimal("0"))
            status = self._parse_status(row.get("状态"), "节点基础信息", row_no, "状态", errors, default="enabled")
            version = self._parse_required_or_default_text(row.get("版本"), "V1")
            sort_order = self._parse_int(row.get("排序"), "节点基础信息", row_no, "排序", errors, default=0)

            if code:
                if code in seen_codes:
                    self._add_error(errors, "节点基础信息", row_no, "节点编码", "同一导入文件中编码重复")
                else:
                    seen_codes.add(code)
                if node_repo.get_by_code(code):
                    self._add_error(errors, "节点基础信息", row_no, "节点编码", "编码已存在")
            if not all(value is not None for value in (code, name, node_type, staff, area, status, version, sort_order)):
                continue

            parsed_nodes[code] = ParsedNodeRow(
                row_no=row_no,
                code=code,
                data={
                    "code": code,
                    "name": name,
                    "node_type": node_type,
                    "staff": staff,
                    "area": area,
                    "description": self._optional_text(row.get("描述")),
                    "status": status,
                    "version": version,
                    "sort_order": sort_order,
                    "remark": self._optional_text(row.get("备注")),
                },
                material_inputs=[],
                consumables=[],
                public_services=[],
                equipment=[],
                outputs=[],
            )

        self._append_node_reference_rows(
            parsed_nodes,
            material_rows,
            "节点输入原料",
            "原料编码",
            "节点编码",
            errors,
            lambda row, row_no, node_code: self._build_node_material_input(
                row,
                row_no,
                node_code,
                errors,
                material_repo,
                parsed_nodes[node_code].data["status"],
            ),
            lambda parsed, payload: parsed.material_inputs.append(payload),
        )
        self._append_node_reference_rows(
            parsed_nodes,
            consumable_rows,
            "节点消耗品",
            "消耗品编码",
            "节点编码",
            errors,
            lambda row, row_no, node_code: self._build_node_consumable(
                row,
                row_no,
                node_code,
                errors,
                consumable_repo,
                parsed_nodes[node_code].data["status"],
            ),
            lambda parsed, payload: parsed.consumables.append(payload),
        )
        self._append_node_reference_rows(
            parsed_nodes,
            public_service_rows,
            "节点公共服务",
            "公共服务编码",
            "节点编码",
            errors,
            lambda row, row_no, node_code: self._build_node_public_service(
                row,
                row_no,
                node_code,
                errors,
                public_service_repo,
                parsed_nodes[node_code].data["status"],
            ),
            lambda parsed, payload: parsed.public_services.append(payload),
        )
        self._append_node_equipment_rows(parsed_nodes, equipment_rows, errors)
        self._append_node_output_rows(parsed_nodes, output_rows, errors, product_repo)

        payloads: list[ProcessNodeCreateWithChildren] = []
        for parsed in parsed_nodes.values():
            if parsed.data["status"] == "enabled" and not parsed.outputs:
                self._add_error(errors, "节点基础信息", parsed.row_no, "状态", "启用节点时至少需要配置一个输出产品")
            try:
                payloads.append(
                    ProcessNodeCreateWithChildren(
                        **parsed.data,
                        material_inputs=parsed.material_inputs,
                        consumables=parsed.consumables,
                        public_services=parsed.public_services,
                        equipment=parsed.equipment,
                        outputs=parsed.outputs,
                    )
                )
            except ValidationError as exc:
                self._append_validation_errors(errors, "节点基础信息", parsed.row_no, exc)

        if errors:
            self._raise_import_errors(errors)

        imported_codes: list[str] = []
        try:
            for payload in payloads:
                self._persist_node(payload, operator)
                imported_codes.append(payload.code)
            self.process_service.system_service.record_operation(
                operator,
                "导入工艺节点库",
                "process_node",
                detail=f"批量导入工艺节点 {len(imported_codes)} 条",
            )
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise
        return ProcessConfigImportResultOut(module=config.module, imported_count=len(imported_codes), imported_codes=imported_codes).model_dump()

    def _import_routes(self, config: ProcessExcelModuleConfig, workbook: Workbook, operator: User) -> dict[str, Any]:
        errors: list[dict[str, Any]] = []
        base_rows = self._sheet_rows(workbook, "路线基础信息", ROUTE_SHEET_HEADERS["路线基础信息"])
        node_rows = self._sheet_rows(workbook, "路线节点链路", ROUTE_SHEET_HEADERS["路线节点链路"])

        route_repo = ProcessRouteRepository(self.db)
        material_repo = self.process_service._repository("material")
        product_repo = self.process_service._repository("product")
        node_repo = ProcessNodeRepository(self.db)

        parsed_routes: dict[str, ParsedRouteRow] = {}
        seen_codes: set[str] = set()

        for row_no, row in base_rows:
            code = self._required_text(row, "路线编码", "路线基础信息", row_no, errors)
            name = self._required_text(row, "路线名称", "路线基础信息", row_no, errors)
            input_material_code = self._required_text(row, "输入原料编码", "路线基础信息", row_no, errors)
            final_product_code = self._required_text(row, "最终产品编码", "路线基础信息", row_no, errors)
            status = self._parse_status(row.get("状态"), "路线基础信息", row_no, "状态", errors, default="enabled")
            sort_order = self._parse_int(row.get("排序"), "路线基础信息", row_no, "排序", errors, default=0)
            version = self._parse_required_or_default_text(row.get("版本"), "V1")

            if code:
                if code in seen_codes:
                    self._add_error(errors, "路线基础信息", row_no, "路线编码", "同一导入文件中编码重复")
                else:
                    seen_codes.add(code)
                if route_repo.get_by_code(code):
                    self._add_error(errors, "路线基础信息", row_no, "路线编码", "编码已存在")

            input_material = self._active_item_by_code(material_repo, input_material_code) if input_material_code else None
            final_product = self._active_item_by_code(product_repo, final_product_code) if final_product_code else None
            if input_material_code and not input_material:
                self._add_error(errors, "路线基础信息", row_no, "输入原料编码", "引用的原料不存在或已删除")
            if final_product_code and not final_product:
                self._add_error(errors, "路线基础信息", row_no, "最终产品编码", "引用的产品不存在或已删除")
            if status == "enabled" and input_material and input_material.status != "enabled":
                self._add_error(errors, "路线基础信息", row_no, "输入原料编码", "启用路线时引用的原料必须为启用状态")
            if status == "enabled" and final_product and final_product.status != "enabled":
                self._add_error(errors, "路线基础信息", row_no, "最终产品编码", "启用路线时引用的产品必须为启用状态")

            if not all(value is not None for value in (code, name, input_material, final_product, status, sort_order, version)):
                continue

            parsed_routes[code] = ParsedRouteRow(
                row_no=row_no,
                code=code,
                data={
                    "code": code,
                    "name": name,
                    "input_material_id": input_material.id,
                    "final_product_id": final_product.id,
                    "version": version,
                    "description": self._optional_text(row.get("描述")),
                    "status": status,
                    "sort_order": sort_order,
                    "remark": self._optional_text(row.get("备注")),
                },
                nodes=[],
            )

        sort_orders_by_route: dict[str, dict[int, int]] = {}
        for row_no, row in node_rows:
            route_code = self._required_text(row, "路线编码", "路线节点链路", row_no, errors)
            node_code = self._required_text(row, "节点编码", "路线节点链路", row_no, errors)
            sort_order = self._parse_int(row.get("排序"), "路线节点链路", row_no, "排序", errors, default=0)
            node_params_json = self._parse_json_text(row.get("节点参数JSON"), "路线节点链路", row_no, "节点参数JSON", errors)
            if not route_code or route_code not in parsed_routes:
                if route_code:
                    self._add_error(errors, "路线节点链路", row_no, "路线编码", "关联的路线编码不存在于“路线基础信息”Sheet")
                continue

            parsed_route = parsed_routes[route_code]
            route_sort_orders = sort_orders_by_route.setdefault(route_code, {})
            if sort_order is not None and sort_order in route_sort_orders:
                self._add_error(errors, "路线节点链路", row_no, "排序", "同一路线下节点顺序不能重复")
            elif sort_order is not None:
                route_sort_orders[sort_order] = row_no

            node = self._active_item_by_code(node_repo, node_code) if node_code else None
            if node_code and not node:
                self._add_error(errors, "路线节点链路", row_no, "节点编码", "引用的工艺节点不存在或已删除")
            if parsed_route.data["status"] == "enabled" and node and node.status != "enabled":
                self._add_error(errors, "路线节点链路", row_no, "节点编码", "启用路线时引用的节点必须为启用状态")

            if node is None or sort_order is None:
                continue

            try:
                parsed_route.nodes.append(
                    ProcessRouteNodePayload(
                        node_id=node.id,
                        sort_order=sort_order,
                        node_params_json=node_params_json,
                        remark=self._optional_text(row.get("备注")),
                    )
                )
            except ValidationError as exc:
                self._append_validation_errors(errors, "路线节点链路", row_no, exc)

        payloads: list[ProcessRouteCreateWithNodes] = []
        for parsed in parsed_routes.values():
            if parsed.data["status"] == "enabled" and not parsed.nodes:
                self._add_error(errors, "路线基础信息", parsed.row_no, "状态", "启用路线时至少需要配置一个节点")
            try:
                payloads.append(ProcessRouteCreateWithNodes(**parsed.data, nodes=parsed.nodes))
            except ValidationError as exc:
                self._append_validation_errors(errors, "路线基础信息", parsed.row_no, exc)

        if errors:
            self._raise_import_errors(errors)

        imported_codes: list[str] = []
        try:
            for payload in payloads:
                self._persist_route(payload, operator)
                imported_codes.append(payload.code)
            self.process_service.system_service.record_operation(
                operator,
                "导入工艺路线库",
                "process_route",
                detail=f"批量导入工艺路线 {len(imported_codes)} 条",
            )
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise
        return ProcessConfigImportResultOut(module=config.module, imported_count=len(imported_codes), imported_codes=imported_codes).model_dump()

    def _append_node_reference_rows(
        self,
        parsed_nodes: dict[str, ParsedNodeRow],
        rows: list[tuple[int, dict[str, Any]]],
        sheet_name: str,
        code_field: str,
        node_field: str,
        errors: list[dict[str, Any]],
        builder: Any,
        append_child: Any,
    ) -> None:
        for row_no, row in rows:
            node_code = self._required_text(row, node_field, sheet_name, row_no, errors)
            if not node_code or node_code not in parsed_nodes:
                if node_code:
                    self._add_error(errors, sheet_name, row_no, node_field, "关联的节点编码不存在于“节点基础信息”Sheet")
                continue
            payload = builder(row, row_no, node_code)
            if payload is not None:
                append_child(parsed_nodes[node_code], payload)

    def _append_node_equipment_rows(
        self,
        parsed_nodes: dict[str, ParsedNodeRow],
        rows: list[tuple[int, dict[str, Any]]],
        errors: list[dict[str, Any]],
    ) -> None:
        for row_no, row in rows:
            node_code = self._required_text(row, "节点编码", "节点设备投资", row_no, errors)
            if not node_code or node_code not in parsed_nodes:
                if node_code:
                    self._add_error(errors, "节点设备投资", row_no, "节点编码", "关联的节点编码不存在于“节点基础信息”Sheet")
                continue
            equipment_name = self._required_text(row, "设备名称", "节点设备投资", row_no, errors)
            quantity = self._parse_decimal(row.get("数量"), "节点设备投资", row_no, "数量", errors, default=Decimal("0"))
            sort_order = self._parse_int(row.get("排序"), "节点设备投资", row_no, "排序", errors, default=0)
            if not all(value is not None for value in (equipment_name, quantity, sort_order)):
                continue
            try:
                parsed_nodes[node_code].equipment.append(
                    ProcessNodeEquipmentPayload(
                        equipment_name=equipment_name,
                        equipment_type=self._optional_text(row.get("设备类型")),
                        quantity=quantity,
                        currency=self._required_or_default_text(row.get("币种"), "CNY"),
                        sort_order=sort_order,
                        remark=self._optional_text(row.get("备注")),
                    )
                )
            except ValidationError as exc:
                self._append_validation_errors(errors, "节点设备投资", row_no, exc)

    def _append_node_output_rows(
        self,
        parsed_nodes: dict[str, ParsedNodeRow],
        rows: list[tuple[int, dict[str, Any]]],
        errors: list[dict[str, Any]],
        product_repo: Any,
    ) -> None:
        for row_no, row in rows:
            node_code = self._required_text(row, "节点编码", "节点输出产品", row_no, errors)
            if not node_code or node_code not in parsed_nodes:
                if node_code:
                    self._add_error(errors, "节点输出产品", row_no, "节点编码", "关联的节点编码不存在于“节点基础信息”Sheet")
                continue
            product_code = self._required_text(row, "产品编码", "节点输出产品", row_no, errors)
            product = self._active_item_by_code(product_repo, product_code) if product_code else None
            if product_code and not product:
                self._add_error(errors, "节点输出产品", row_no, "产品编码", "引用的产品不存在或已删除")
            if parsed_nodes[node_code].data["status"] == "enabled" and product and product.status != "enabled":
                self._add_error(errors, "节点输出产品", row_no, "产品编码", "启用节点时引用的产品必须为启用状态")
            output_per_ton = self._parse_decimal(row.get("产出量/吨"), "节点输出产品", row_no, "产出量/吨", errors, default=Decimal("0"))
            unit = self._required_text(row, "单位", "节点输出产品", row_no, errors)
            is_main_product = self._parse_bool(row.get("是否主产品"), "节点输出产品", row_no, "是否主产品", errors, default=False)
            sort_order = self._parse_int(row.get("排序"), "节点输出产品", row_no, "排序", errors, default=0)
            if not all(value is not None for value in (product, output_per_ton, unit, is_main_product, sort_order)):
                continue
            try:
                parsed_nodes[node_code].outputs.append(
                    ProcessNodeOutputPayload(
                        product_id=product.id,
                        output_per_ton=output_per_ton,
                        unit=unit,
                        is_main_product=is_main_product,
                        sort_order=sort_order,
                        remark=self._optional_text(row.get("备注")),
                    )
                )
            except ValidationError as exc:
                self._append_validation_errors(errors, "节点输出产品", row_no, exc)

    def _build_node_material_input(
        self,
        row: dict[str, Any],
        row_no: int,
        node_code: str,
        errors: list[dict[str, Any]],
        material_repo: Any,
        node_status: str,
    ) -> ProcessNodeMaterialInputPayload | None:
        material_code = self._required_text(row, "原料编码", "节点输入原料", row_no, errors)
        material = self._active_item_by_code(material_repo, material_code) if material_code else None
        if material_code and not material:
            self._add_error(errors, "节点输入原料", row_no, "原料编码", "引用的原料不存在或已删除")
        if node_status == "enabled" and material and material.status != "enabled":
            self._add_error(errors, "节点输入原料", row_no, "原料编码", "启用节点时引用的原料必须为启用状态")
        amount_per_ton = self._parse_decimal(row.get("投入量/吨"), "节点输入原料", row_no, "投入量/吨", errors, default=Decimal("0"))
        unit = self._required_text(row, "单位", "节点输入原料", row_no, errors)
        sort_order = self._parse_int(row.get("排序"), "节点输入原料", row_no, "排序", errors, default=0)
        if not all(value is not None for value in (material, amount_per_ton, unit, sort_order)):
            return None
        try:
            return ProcessNodeMaterialInputPayload(
                material_id=material.id,
                amount_per_ton=amount_per_ton,
                unit=unit,
                sort_order=sort_order,
                remark=self._optional_text(row.get("备注")),
            )
        except ValidationError as exc:
            self._append_validation_errors(errors, "节点输入原料", row_no, exc)
            return None

    def _build_node_consumable(
        self,
        row: dict[str, Any],
        row_no: int,
        node_code: str,
        errors: list[dict[str, Any]],
        consumable_repo: Any,
        node_status: str,
    ) -> ProcessNodeConsumablePayload | None:
        consumable_code = self._required_text(row, "消耗品编码", "节点消耗品", row_no, errors)
        consumable = self._active_item_by_code(consumable_repo, consumable_code) if consumable_code else None
        if consumable_code and not consumable:
            self._add_error(errors, "节点消耗品", row_no, "消耗品编码", "引用的消耗品不存在或已删除")
        if node_status == "enabled" and consumable and consumable.status != "enabled":
            self._add_error(errors, "节点消耗品", row_no, "消耗品编码", "启用节点时引用的消耗品必须为启用状态")
        amount_per_ton = self._parse_decimal(row.get("消耗量/吨"), "节点消耗品", row_no, "消耗量/吨", errors, default=Decimal("0"))
        unit = self._required_text(row, "单位", "节点消耗品", row_no, errors)
        sort_order = self._parse_int(row.get("排序"), "节点消耗品", row_no, "排序", errors, default=0)
        if not all(value is not None for value in (consumable, amount_per_ton, unit, sort_order)):
            return None
        try:
            return ProcessNodeConsumablePayload(
                consumable_id=consumable.id,
                amount_per_ton=amount_per_ton,
                unit=unit,
                sort_order=sort_order,
                remark=self._optional_text(row.get("备注")),
            )
        except ValidationError as exc:
            self._append_validation_errors(errors, "节点消耗品", row_no, exc)
            return None

    def _build_node_public_service(
        self,
        row: dict[str, Any],
        row_no: int,
        node_code: str,
        errors: list[dict[str, Any]],
        public_service_repo: Any,
        node_status: str,
    ) -> ProcessNodePublicServicePayload | None:
        service_code = self._required_text(row, "公共服务编码", "节点公共服务", row_no, errors)
        public_service = self._active_item_by_code(public_service_repo, service_code) if service_code else None
        if service_code and not public_service:
            self._add_error(errors, "节点公共服务", row_no, "公共服务编码", "引用的公共服务不存在或已删除")
        if node_status == "enabled" and public_service and public_service.status != "enabled":
            self._add_error(errors, "节点公共服务", row_no, "公共服务编码", "启用节点时引用的公共服务必须为启用状态")
        amount_per_ton = self._parse_decimal(row.get("消耗量/吨"), "节点公共服务", row_no, "消耗量/吨", errors, default=Decimal("0"))
        unit = self._required_text(row, "单位", "节点公共服务", row_no, errors)
        sort_order = self._parse_int(row.get("排序"), "节点公共服务", row_no, "排序", errors, default=0)
        if not all(value is not None for value in (public_service, amount_per_ton, unit, sort_order)):
            return None
        try:
            return ProcessNodePublicServicePayload(
                public_service_id=public_service.id,
                amount_per_ton=amount_per_ton,
                unit=unit,
                sort_order=sort_order,
                remark=self._optional_text(row.get("备注")),
            )
        except ValidationError as exc:
            self._append_validation_errors(errors, "节点公共服务", row_no, exc)
            return None

    def _persist_library(self, kind: ProcessLibraryKind, payload: ProcessLibraryCreateWithPrices, operator: User) -> None:
        config = self.process_service._config(kind)
        repo = self.process_service._repository(kind)
        self.process_service._validate_code_unique(repo, payload.code)

        item_data = payload.model_dump(exclude={"region_prices"})
        item = config.model(**item_data, created_by=operator.id, updated_by=operator.id, is_deleted=False)
        repo.add(item)
        normalized_prices = self.process_service._normalize_region_prices(payload.region_prices, item.unit, allow_default=True)
        self.process_service._sync_region_prices(repo, item, normalized_prices, operator)

    def _persist_node(self, payload: ProcessNodeCreateWithChildren, operator: User) -> None:
        repo = ProcessNodeRepository(self.db)
        self.process_service._validate_node_code_unique(repo, payload.code)
        self.process_service._validate_node_payload(payload.status, payload)
        node_data = payload.model_dump(
            exclude={"material_inputs", "consumables", "public_services", "equipment", "labor", "outputs"}
        )
        from app.models.process_config import ProcessNode  # local import to keep Excel service dependency narrow

        node = ProcessNode(**node_data, created_by=operator.id, updated_by=operator.id, is_deleted=False)
        repo.add(node)
        self.process_service._replace_node_children(repo, node.id, payload)

    def _persist_route(self, payload: ProcessRouteCreateWithNodes, operator: User) -> None:
        repo = ProcessRouteRepository(self.db)
        self.process_service._validate_route_code_unique(repo, payload.code)
        self.process_service._validate_route_payload(
            payload.status,
            payload.input_material_id,
            payload.final_product_id,
            payload.nodes,
        )
        route_data = payload.model_dump(exclude={"nodes"})
        from app.models.process_config import ProcessRoute  # local import to keep Excel service dependency narrow

        route = ProcessRoute(**route_data, created_by=operator.id, updated_by=operator.id, is_deleted=False)
        repo.add(route)
        repo.replace_nodes(route.id, [self.process_service._route_node_payload_dump(row) for row in payload.nodes])

    def _module_config(self, module: ProcessExcelModule) -> ProcessExcelModuleConfig:
        config = MODULE_CONFIGS.get(module)
        if not config:
            raise AppException("不支持的工艺配置模块", status_code=404, code=404)
        return config

    def _create_workbook(self, sheet_name: str, headers: list[str]) -> Workbook:
        workbook = Workbook()
        workbook.active.title = sheet_name
        self._prepare_header(workbook.active, headers)
        return workbook

    def _create_sheet(self, workbook: Workbook, sheet_name: str, headers: list[str]):
        sheet = workbook.create_sheet(sheet_name)
        self._prepare_header(sheet, headers)
        return sheet

    def _prepare_header(self, sheet: Any, headers: list[str]) -> None:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

    def _finalize_sheet(self, sheet: Any) -> None:
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            width = max(12, min(max((len(value) for value in values), default=12) + 2, 40))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    def _append_instruction_sheet(self, workbook: Workbook, rows: list[tuple[str, str]]) -> None:
        sheet = workbook.create_sheet("填写说明")
        self._prepare_header(sheet, ["字段", "说明"])
        for field_name, description in rows:
            sheet.append([field_name, description])
        self._finalize_sheet(sheet)

    def _workbook_bytes(self, workbook: Workbook) -> bytes:
        for sheet in workbook.worksheets:
            self._finalize_sheet(sheet)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _load_workbook(self, content: bytes) -> Workbook:
        try:
            return load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise AppException("仅支持导入合法的 .xlsx 文件") from exc

    def _sheet_rows(self, workbook: Workbook, sheet_name: str, headers: list[str]) -> list[tuple[int, dict[str, Any]]]:
        errors: list[dict[str, Any]] = []
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if sheet is None:
            self._raise_import_errors([self._error(sheet_name, 1, "Sheet", "缺少必填 Sheet")])
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            self._raise_import_errors([self._error(sheet_name, 1, "表头", "Sheet 为空，缺少表头")])
        actual_headers = [self._optional_text(value) or "" for value in values[0]]
        header_index: dict[str, int] = {}
        for index, header in enumerate(actual_headers):
            if header and header not in header_index:
                header_index[header] = index
        missing_headers = [header for header in headers if header not in header_index]
        if missing_headers:
            raise AppException(
                "Excel 导入校验失败",
                data={"errors": [self._error(sheet_name, 1, header, "缺少必填表头") for header in missing_headers]},
            )

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_no, values_row in enumerate(values[1:], start=2):
            row = {
                header: values_row[header_index[header]] if header_index[header] < len(values_row) else None
                for header in headers
            }
            if all(self._is_empty(value) for value in row.values()):
                continue
            rows.append((row_no, row))
        return rows

    def _append_validation_errors(self, errors: list[dict[str, Any]], sheet: str, row: int, exc: ValidationError) -> None:
        for item in exc.errors():
            field = ".".join(str(part) for part in item.get("loc", [])) or "字段"
            errors.append(self._error(sheet, row, field, item.get("msg", "格式错误")))

    def _raise_import_errors(self, errors: list[dict[str, Any]]) -> None:
        ordered_errors = sorted(errors, key=lambda item: (item["sheet"], item["row"], item["field"]))
        raise AppException("Excel 导入校验失败", data={"errors": ordered_errors})

    def _error(self, sheet: str, row: int, field: str, message: str) -> dict[str, Any]:
        return {"sheet": sheet, "row": row, "field": field, "message": message}

    def _add_error(self, errors: list[dict[str, Any]], sheet: str, row: int, field: str, message: str) -> None:
        errors.append(self._error(sheet, row, field, message))

    def _required_text(
        self,
        row: dict[str, Any],
        field: str,
        sheet: str,
        row_no: int,
        errors: list[dict[str, Any]],
    ) -> str | None:
        value = self._optional_text(row.get(field))
        if value is None:
            self._add_error(errors, sheet, row_no, field, "不能为空")
        return value

    def _parse_status(
        self,
        value: Any,
        sheet: str,
        row_no: int,
        field: str,
        errors: list[dict[str, Any]],
        *,
        default: str,
    ) -> str | None:
        text = self._optional_text(value) or default
        status = STATUS_ALIASES.get(text)
        if status is None:
            self._add_error(errors, sheet, row_no, field, "仅支持 启用/草稿/停用 或 enabled/draft/disabled")
        return status

    def _parse_node_type(
        self,
        value: Any,
        sheet: str,
        row_no: int,
        field: str,
        errors: list[dict[str, Any]],
    ) -> str | None:
        text = self._optional_text(value)
        if text is None:
            self._add_error(errors, sheet, row_no, field, "不能为空")
            return None
        node_type = NODE_TYPE_ALIASES.get(text)
        if node_type is None:
            self._add_error(errors, sheet, row_no, field, "仅支持 预处理/湿法冶金/火法冶金/后处理 或对应英文编码")
        return node_type

    def _parse_decimal(
        self,
        value: Any,
        sheet: str,
        row_no: int,
        field: str,
        errors: list[dict[str, Any]],
        *,
        default: Decimal | None = None,
    ) -> Decimal | None:
        if self._is_empty(value):
            return default
        try:
            return Decimal(str(value).strip())
        except (InvalidOperation, ValueError):
            self._add_error(errors, sheet, row_no, field, "必须为合法数字")
            return None

    def _parse_int(
        self,
        value: Any,
        sheet: str,
        row_no: int,
        field: str,
        errors: list[dict[str, Any]],
        *,
        default: int | None = None,
    ) -> int | None:
        if self._is_empty(value):
            return default
        try:
            if isinstance(value, float) and not value.is_integer():
                raise ValueError("not integer")
            return int(str(value).strip())
        except (TypeError, ValueError):
            self._add_error(errors, sheet, row_no, field, "必须为整数")
            return None

    def _parse_bool(
        self,
        value: Any,
        sheet: str,
        row_no: int,
        field: str,
        errors: list[dict[str, Any]],
        *,
        default: bool,
    ) -> bool | None:
        if self._is_empty(value):
            return default
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in BOOL_ALIASES:
            return BOOL_ALIASES[normalized]
        self._add_error(errors, sheet, row_no, field, "仅支持 是/否 或 true/false")
        return None

    def _parse_json_text(
        self,
        value: Any,
        sheet: str,
        row_no: int,
        field: str,
        errors: list[dict[str, Any]],
    ) -> str | None:
        text = self._optional_text(value)
        if text is None:
            return None
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            self._add_error(errors, sheet, row_no, field, "必须为合法 JSON")
            return None
        return json.dumps(parsed, ensure_ascii=False, separators=(",", ":"))

    def _optional_text(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        stripped = str(value).strip()
        return stripped or None

    def _required_or_default_text(self, value: Any, default: str) -> str:
        return self._optional_text(value) or default

    def _parse_required_or_default_text(self, value: Any, default: str) -> str:
        return self._required_or_default_text(value, default)

    def _optional_int(self, value: Any) -> int | None:
        if self._is_empty(value):
            return None
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return None

    def _status_label(self, status: str) -> str:
        return STATUS_LABELS.get(status, status)

    def _node_type_label(self, node_type: str) -> str:
        return NODE_TYPE_LABELS.get(node_type, node_type)

    def _active_item_by_code(self, repo: Any, code: str | None) -> Any | None:
        if not code:
            return None
        item = repo.get_by_code(code)
        if item is None or getattr(item, "is_deleted", False):
            return None
        return item

    def _is_empty(self, value: Any) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())
